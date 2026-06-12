import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QFrame, QMessageBox, QFileDialog, QGridLayout, QScrollArea)
from PySide6.QtCore import Qt
from datetime import datetime
from themes.style import (PRIMARY, PRIMARY_LIGHT, SUCCESS, SUCCESS_LIGHT, DANGER, DANGER_LIGHT,
    WARNING, WARNING_LIGHT, INFO, INFO_LIGHT, PURPLE, PURPLE_LIGHT, TEAL, TEAL_LIGHT,
    PINK, PINK_LIGHT, NAN_COLOR, NAN_TEXT, BORDER, TEXT_MAIN, TEXT_SUB)

BTN_SEC = 'QPushButton { background: #F3F4F6; color: #374151; border: 1px solid #E5E7EB; border-radius: 8px; padding: 8px 18px; font-weight: 500; font-size: 12px; } QPushButton:hover { background: #E5E7EB; }'


def report_card(icon, title, description, accent, light, callback):
    card = QFrame()
    card.setStyleSheet(f'QFrame {{ background: white; border: 1px solid {BORDER}; border-radius: 14px; }} QFrame:hover {{ border-color: {accent}; background: {light}; }}')
    card.setMinimumHeight(110); card.setCursor(Qt.PointingHandCursor)
    layout = QVBoxLayout(card); layout.setContentsMargins(18, 14, 18, 14); layout.setSpacing(8)
    top = QHBoxLayout()
    pill = QLabel(icon); pill.setFixedSize(38, 38); pill.setAlignment(Qt.AlignCenter)
    pill.setStyleSheet(f'background: {light}; border-radius: 10px; font-size: 19px;')
    top.addWidget(pill); top.addStretch()
    arrow = QLabel('→'); arrow.setStyleSheet(f'color: {accent}; font-size: 18px; font-weight: bold; background: transparent;')
    top.addWidget(arrow); layout.addLayout(top)
    t = QLabel(title); t.setStyleSheet(f'color: {TEXT_MAIN}; font-size: 13px; font-weight: 700; background: transparent;')
    d = QLabel(description); d.setStyleSheet(f'color: {TEXT_SUB}; font-size: 11px; background: transparent;'); d.setWordWrap(True)
    layout.addWidget(t); layout.addWidget(d)
    card.mousePressEvent = lambda e: callback()
    return card


class ReportsWidget(QWidget):
    def __init__(self, session):
        super().__init__()
        self.session = session
        self.setStyleSheet('background: transparent;')
        self._setup_ui()

    def _setup_ui(self):
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setStyleSheet('QScrollArea { border: none; background: transparent; }')
        container = QWidget(); container.setStyleSheet('background: transparent;')
        layout = QVBoxLayout(container); layout.setContentsMargins(28, 24, 28, 28); layout.setSpacing(22)

        # Header
        year = datetime.now().year
        hrow = QHBoxLayout()
        col = QVBoxLayout(); col.setSpacing(3)
        col.addWidget(QLabel('Rapports & Exports') if False else self._h('Rapports & Exports', 20, '#1A1D2E'))
        col.addWidget(self._h('Générez et exportez vos rapports en PDF, Excel ou CSV', 13, '#6B7280'))
        hrow.addLayout(col); hrow.addStretch()
        self.year_combo = QComboBox()
        self.year_combo.addItems([str(y) for y in range(year, year-5, -1)])
        self.year_combo.setFixedHeight(40)
        self.year_combo.setStyleSheet(f'QComboBox {{ background: white; border: 1.5px solid {BORDER}; border-radius: 10px; color: {TEXT_MAIN}; padding: 0 14px; font-size: 13px; min-width: 110px; }} QComboBox QAbstractItemView {{ background: white; border: 1.5px solid #C7D2FE; border-radius: 10px; color: {TEXT_MAIN}; padding: 4px; outline: none; }} QComboBox QAbstractItemView::item {{ padding: 9px 14px; border-radius: 6px; margin: 1px 4px; color: {TEXT_MAIN}; }} QComboBox QAbstractItemView::item:hover {{ background: #F5F3FF; color: {PRIMARY}; }}')
        yr_lbl = QLabel('Année :'); yr_lbl.setStyleSheet(f'color: {TEXT_SUB}; font-size: 13px; background: transparent;')
        hrow.addWidget(yr_lbl); hrow.addWidget(self.year_combo)
        layout.addLayout(hrow)

        # PDF section
        layout.addWidget(self._section_lbl('📄  Rapports PDF'))
        grid1 = QGridLayout(); grid1.setSpacing(14)
        pdfs = [
            ('📋', 'Liste des Élèves',    'Export PDF complet avec NAN, transport, re-inscription',    PRIMARY, PRIMARY_LIGHT, self._students_pdf),
            ('💰', 'Rapport Financier',   'Revenus, dépenses, paiements par mois (assurance séparée)', SUCCESS, SUCCESS_LIGHT, self._financial_pdf),
            ('🛡️', 'Rapport Assurances',  'Statut assurance de tous les élèves (payée / non payée)',   TEAL,    TEAL_LIGHT,    self._insurance_pdf),
            ('👥', 'Liste du Personnel',  'Enseignants, staff, chauffeurs',                            INFO,    INFO_LIGHT,    self._employees_pdf),
            ('🚌', 'Rapport Transport',   'Abonnés, montants, flotte',                                 '#0369A1','#DBEAFE',   self._transport_pdf),
        ]
        for i, args in enumerate(pdfs): grid1.addWidget(report_card(*args), i // 2, i % 2)
        layout.addLayout(grid1)

        # Excel section
        layout.addWidget(self._section_lbl('📊  Exports Excel'))
        grid2 = QGridLayout(); grid2.setSpacing(14)
        excels = [
            ('💳', 'Paiements Complets',  'Tous les paiements avec mois payés/impayés/NAN',      PURPLE, PURPLE_LIGHT, self._payments_excel),
            ('🛡️', 'Assurances Excel',    'Statut et historique assurance par élève',             TEAL,   TEAL_LIGHT,  self._insurance_excel),
            ('📈', 'Analytiques Mensuels','Répartition revenus + graphiques',                     WARNING,WARNING_LIGHT,self._analytics_excel),
            ('🔄', 'Re-inscriptions',     'Statut re-inscription de tous les élèves',             PINK,   PINK_LIGHT,  self._reinscription_excel),
            ('⊘',  'Rapport Mois NAN',   'Tous les mois NAN et leur impact',                     '#854D0E', NAN_COLOR, self._nan_report_excel),
        ]
        for i, args in enumerate(excels): grid2.addWidget(report_card(*args), i // 2, i % 2)
        layout.addLayout(grid2)

        # Backup section
        layout.addWidget(self._section_lbl('⚙️  Système'))
        grid3 = QGridLayout(); grid3.setSpacing(14)
        grid3.addWidget(report_card('💾', 'Sauvegarde Base de Données', 'Copie SQLite locale', '#6D28D9', '#EDE9FE', self._backup_db), 0, 0)
        grid3.addWidget(QFrame(), 0, 1)
        layout.addLayout(grid3)

        self.status_lbl = QLabel()
        self.status_lbl.setStyleSheet(f'color: {SUCCESS}; font-size: 12px; font-weight: 500; background: transparent;')
        layout.addWidget(self.status_lbl)
        layout.addStretch()

        scroll.setWidget(container)
        ol = QVBoxLayout(self); ol.setContentsMargins(0,0,0,0); ol.addWidget(scroll)

    def _h(self, text, size, color):
        l = QLabel(text); l.setStyleSheet(f'color: {color}; font-size: {size}px; font-weight: {"800" if size > 14 else "400"}; background: transparent;')
        return l

    def _section_lbl(self, text):
        l = QLabel(text); l.setStyleSheet(f'color: {TEXT_MAIN}; font-size: 14px; font-weight: 700; background: transparent;')
        return l

    def _ok(self, path): self.status_lbl.setText(f'✅  Export réussi : {path}')

    def _students_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors; from reportlab.lib.units import mm
            from models.database import Student, MonthRecord
            from themes.style import REINSCRIPTION_LABELS, SCHOOL_MONTHS

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'eleves_{datetime.now().year}.pdf', 'PDF (*.pdf)')
            if not path: return
            students = self.session.query(Student).filter_by(active=True).order_by(Student.class_name, Student.last_name).all()
            doc = SimpleDocTemplate(path, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
            story = []
            story.append(Paragraph(f'<b>Le Schéma — Liste des Élèves {datetime.now().year}</b>',
                ParagraphStyle('t', fontSize=14, textColor=colors.HexColor('#4F46E5'), alignment=1, spaceAfter=4)))
            story.append(Paragraph(f'{len(students)} élèves — Généré le {datetime.now().strftime("%d/%m/%Y")}',
                ParagraphStyle('s', fontSize=9, textColor=colors.HexColor('#6B7280'), alignment=1, spaceAfter=6)))
            story.append(Spacer(1, 5*mm))

            data = [['Code','Nom','Prénom','Classe','Parent','Tél.','Mensuel','Transport','Assurance','Re-insc.']]
            for s in students:
                transport_str = f'{getattr(s,"transport_fee",0):.0f} MAD' if getattr(s,'has_transport',False) else 'NA'
                reinsc = REINSCRIPTION_LABELS.get(getattr(s,'reinscription_status','pending'), '—')
                data.append([s.code or '', s.last_name or '', s.first_name or '', s.class_name or '',
                              s.parent_name or '', s.parent_phone or '',
                              f'{s.monthly_fee:.0f} MAD', transport_str,
                              '✓ Payée' if s.insurance_paid else '✗ Non',
                              reinsc])
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#4F46E5')), ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'), ('FONTSIZE',(0,0),(-1,-1),8),
                ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#EAEDF3')),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F9FAFB')]),
                ('PADDING',(0,0),(-1,-1),5),
            ]))
            story.append(t); doc.build(story)
            self._ok(path)
        except Exception as e: QMessageBox.critical(self, 'Erreur', str(e))

    def _financial_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors; from reportlab.lib.units import mm
            from models.database import Payment, ExpensePayment, ExpenseCategory
            from themes.style import SCHOOL_MONTHS

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'rapport_financier_{datetime.now().year}.pdf', 'PDF (*.pdf)')
            if not path: return
            year = int(self.year_combo.currentText())
            payments = self.session.query(Payment).filter(Payment.year == year).all()
            # Use ExpensePayment (v4) for expense totals
            expense_payments = self.session.query(ExpensePayment).filter_by(year=year).all()

            # FINANCIAL RULE: insurance excluded from operational revenue
            operational_payments = [p for p in payments if p.payment_type in ('monthly', 'transport')]
            insurance_payments   = [p for p in payments if p.payment_type == 'insurance']
            total_rev   = sum(p.amount for p in operational_payments)  # excludes insurance
            total_ins   = sum(p.amount for p in insurance_payments)
            total_exp   = sum(ep.amount or 0 for ep in expense_payments)
            profit      = total_rev - total_exp  # insurance never enters profit formula

            doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=15*mm, bottomMargin=15*mm)
            story = []
            story.append(Paragraph(f'<b>Rapport Financier {year} — Le Schéma</b>',
                ParagraphStyle('t', fontSize=18, textColor=colors.HexColor('#4F46E5'), alignment=1, spaceAfter=6)))
            story.append(Spacer(1, 8*mm))
            summary = [
                ['Revenus opérationnels (mensualités + transport)', f'{total_rev:,.2f} MAD'],
                ['Dépenses totales',                                 f'{total_exp:,.2f} MAD'],
                ['Bénéfice net (hors assurance)',                    f'{profit:,.2f} MAD'],
                ['Assurances encaissées (suivi séparé)',             f'{total_ins:,.2f} MAD'],
            ]
            st = Table(summary, colWidths=[120*mm, 60*mm])
            st.setStyle(TableStyle([
                ('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'), ('FONTSIZE',(0,0),(-1,-1),11),
                ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#EAEDF3')), ('PADDING',(0,0),(-1,-1),10),
                ('ALIGN',(1,0),(1,-1),'RIGHT'),
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#D1FAE5')),
                ('BACKGROUND',(0,1),(-1,1),colors.HexColor('#FEE2E2')),
                ('BACKGROUND',(0,2),(-1,2),colors.HexColor('#D1FAE5') if profit>=0 else colors.HexColor('#FEE2E2')),
                ('BACKGROUND',(0,3),(-1,3),colors.HexColor('#CCFBF1')),
                ('TEXTCOLOR',(0,3),(-1,3),colors.HexColor('#0F766E')),
            ]))
            story.append(st); story.append(Spacer(1, 8*mm))

            month_data = [['Mois', 'Mensualités', 'Transport', 'Total opérationnel', 'Assurance (séparé)']]
            for m in SCHOOL_MONTHS:
                mo  = sum(p.amount for p in payments if p.month==m and p.payment_type=='monthly')
                tr  = sum(p.amount for p in payments if p.month==m and p.payment_type=='transport')
                ins = sum(p.amount for p in payments if p.month==m and p.payment_type=='insurance')
                total_op = mo + tr
                if total_op + ins > 0:
                    month_data.append([m, f'{mo:.0f}', f'{tr:.0f}', f'{total_op:.0f} MAD', f'{ins:.0f} MAD'])
            if len(month_data) > 1:
                mt = Table(month_data, repeatRows=1)
                mt.setStyle(TableStyle([
                    ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#4F46E5')), ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                    ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'), ('FONTSIZE',(0,0),(-1,-1),9),
                    ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#EAEDF3')),
                    ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F9FAFB')]),
                    ('PADDING',(0,0),(-1,-1),5),
                    # Insurance column tinted
                    ('TEXTCOLOR',(4,1),(4,-1),colors.HexColor('#0F766E')),
                ]))
                story.append(mt)
            doc.build(story); self._ok(path)
        except Exception as e: QMessageBox.critical(self, 'Erreur', str(e))

    def _insurance_pdf(self):
        """PDF report: insurance status for every student — paid / unpaid."""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from models.database import Student, Payment, Setting

            path, _ = QFileDialog.getSaveFileName(
                self, 'Enregistrer', f'rapport_assurances_{datetime.now().year}.pdf', 'PDF (*.pdf)')
            if not path: return

            s_yr = self.session.query(Setting).filter_by(key='school_year').first()
            school_year = s_yr.value if s_yr else '2024-25'

            students = (self.session.query(Student)
                        .filter_by(active=True)
                        .order_by(Student.class_name, Student.last_name).all())

            # Build insurance payment map for this school year
            ins_map = {
                p.student_id: p
                for p in self.session.query(Payment).filter_by(
                    payment_type='insurance', school_year=school_year
                ).all()
            }

            n_paid   = sum(1 for s in students if s.id in ins_map)
            n_unpaid = len(students) - n_paid
            total    = sum(p.amount or 0 for p in ins_map.values())

            doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                    rightMargin=12*mm, leftMargin=12*mm,
                                    topMargin=12*mm, bottomMargin=12*mm)
            story = []
            title_s = ParagraphStyle('t', fontSize=16,
                                     textColor=colors.HexColor('#0F766E'),
                                     alignment=1, fontName='Helvetica-Bold', spaceAfter=4)
            sub_s   = ParagraphStyle('s', fontSize=10,
                                     textColor=colors.HexColor('#6B7280'),
                                     alignment=1, spaceAfter=8)
            story.append(Paragraph(f'Rapport Assurances Scolaires — {school_year}', title_s))
            story.append(Paragraph(
                f'{len(students)} élèves  •  {n_paid} assurés  •  '
                f'{n_unpaid} non assurés  •  Encaissé: {total:,.0f} MAD  •  '
                f'Généré le {datetime.now().strftime("%d/%m/%Y")}', sub_s))
            story.append(Spacer(1, 4*mm))

            data = [['Classe', 'Nom', 'Prénom', 'Montant', 'Statut', 'Date paiement', 'N° Reçu']]
            for s in students:
                p    = ins_map.get(s.id)
                paid = p is not None
                date_str = p.payment_date.strftime('%d/%m/%Y') if (p and p.payment_date) else '—'
                rec_str  = p.receipt_number or '—' if p else '—'
                data.append([
                    s.class_name or '—',
                    s.last_name   or '—',
                    s.first_name  or '—',
                    f'{s.insurance_amount or 500:.0f} MAD',
                    '✓ Payée'    if paid else '✗ Non payée',
                    date_str,
                    rec_str,
                ])

            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F766E')),
                ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
                ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0,0), (-1,-1), 9),
                ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#CCFBF1')),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F0FDFA')]),
                ('PADDING',    (0,0), (-1,-1), 6),
                # Color status column
                ('TEXTCOLOR',  (4,1), (4,-1), colors.HexColor('#059669')),
            ]))
            # Highlight unpaid rows in light red
            for i, s in enumerate(students, start=1):
                if s.id not in ins_map:
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0,i), (-1,i), colors.HexColor('#FEF2F2')),
                        ('TEXTCOLOR',  (4,i), (4,i),  colors.HexColor('#DC2626')),
                    ]))
            story.append(t)
            doc.build(story)
            self._ok(path)
        except Exception as e:
            QMessageBox.critical(self, 'Erreur', str(e))

    def _insurance_excel(self):
        """Excel export: insurance status + history for all students."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from models.database import Student, Payment, Setting

            path, _ = QFileDialog.getSaveFileName(
                self, 'Enregistrer', f'assurances_{datetime.now().year}.xlsx', 'Excel (*.xlsx)')
            if not path: return

            s_yr = self.session.query(Setting).filter_by(key='school_year').first()
            school_year = s_yr.value if s_yr else '2024-25'

            students = (self.session.query(Student)
                        .filter_by(active=True)
                        .order_by(Student.class_name, Student.last_name).all())
            ins_map = {
                p.student_id: p
                for p in self.session.query(Payment).filter_by(
                    payment_type='insurance', school_year=school_year
                ).all()
            }

            wb  = openpyxl.Workbook()

            # ── Sheet 1: Status par élève ────────────────────────────────────
            ws1 = wb.active
            ws1.title = 'Statut Assurances'

            hdr_fill  = PatternFill('solid', fgColor='0F766E')
            paid_fill = PatternFill('solid', fgColor='D1FAE5')
            unpd_fill = PatternFill('solid', fgColor='FEE2E2')
            hdr_font  = Font(bold=True, color='FFFFFF', size=10)
            bd = Border(
                left=Side(style='thin', color='CCFBF1'),
                right=Side(style='thin', color='CCFBF1'),
                top=Side(style='thin', color='CCFBF1'),
                bottom=Side(style='thin', color='CCFBF1'),
            )
            headers = ['Classe', 'Nom', 'Prénom', 'Montant assurance',
                       'Statut', 'Date paiement', 'N° Reçu', 'Année scolaire']
            for col, h in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=h)
                cell.font  = hdr_font
                cell.fill  = hdr_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = bd

            for row_i, s in enumerate(students, start=2):
                p    = ins_map.get(s.id)
                paid = p is not None
                fill = paid_fill if paid else unpd_fill
                vals = [
                    s.class_name or '',
                    s.last_name   or '',
                    s.first_name  or '',
                    s.insurance_amount or 500,
                    'Payée' if paid else 'Non payée',
                    p.payment_date.strftime('%d/%m/%Y') if (p and p.payment_date) else '',
                    p.receipt_number or '' if p else '',
                    school_year,
                ]
                for col_i, v in enumerate(vals, 1):
                    cell = ws1.cell(row=row_i, column=col_i, value=v)
                    cell.fill   = fill
                    cell.border = bd
                    cell.alignment = Alignment(horizontal='center' if col_i > 3 else 'left')

            # Auto-width
            for col in ws1.columns:
                max_len = max(len(str(c.value or '')) for c in col) + 3
                ws1.column_dimensions[get_column_letter(col[0].column)].width = min(max_len, 30)
            ws1.row_dimensions[1].height = 22

            # ── Sheet 2: Summary ─────────────────────────────────────────────
            ws2 = wb.create_sheet('Résumé')
            n_paid   = sum(1 for s in students if s.id in ins_map)
            n_unpaid = len(students) - n_paid
            total_am = sum(p.amount or 0 for p in ins_map.values())

            sum_data = [
                ('Année scolaire',           school_year),
                ('Total élèves',             len(students)),
                ('Assurances payées',        n_paid),
                ('Non assurés',              n_unpaid),
                ('Taux de couverture',       f'{100*n_paid//len(students) if students else 0}%'),
                ('Total encaissé',           f'{total_am:,.2f} MAD'),
                ('Note',                     'Assurance exclue du bénéfice mensuel'),
            ]
            ws2['A1'] = 'RÉSUMÉ ASSURANCES SCOLAIRES'
            ws2['A1'].font = Font(bold=True, size=14, color='0F766E')
            ws2['A1'].alignment = Alignment(horizontal='center')
            ws2.merge_cells('A1:B1')
            for i, (label, value) in enumerate(sum_data, start=2):
                ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
                ws2.cell(row=i, column=2, value=value)
            ws2.column_dimensions['A'].width = 28
            ws2.column_dimensions['B'].width = 22

            # ── Sheet 3: History all insurance payments ───────────────────────
            ws3 = wb.create_sheet('Historique Paiements')
            all_ins = (self.session.query(Payment)
                       .filter_by(payment_type='insurance')
                       .order_by(Payment.payment_date.desc()).all())
            hist_hdrs = ['N° Reçu', 'Élève', 'Classe', 'Montant',
                         'Année scolaire', 'Date paiement']
            for col, h in enumerate(hist_hdrs, 1):
                cell = ws3.cell(row=1, column=col, value=h)
                cell.font  = hdr_font
                cell.fill  = hdr_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = bd
            for row_i, p in enumerate(all_ins, start=2):
                student = self.session.query(Student).filter_by(id=p.student_id).first()
                name = f'{student.first_name} {student.last_name}' if student else '—'
                cls  = student.class_name if student else '—'
                row_vals = [
                    p.receipt_number or '—',
                    name,
                    cls,
                    p.amount or 0,
                    p.school_year or '—',
                    p.payment_date.strftime('%d/%m/%Y') if p.payment_date else '—',
                ]
                for col_i, v in enumerate(row_vals, 1):
                    ws3.cell(row=row_i, column=col_i, value=v).border = bd
            for col in ws3.columns:
                max_len = max(len(str(c.value or '')) for c in col) + 3
                ws3.column_dimensions[get_column_letter(col[0].column)].width = min(max_len, 30)

            wb.save(path)
            self._ok(path)
        except Exception as e:
            QMessageBox.critical(self, 'Erreur', str(e))

    def _employees_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib import colors; from reportlab.lib.units import mm
            from models.database import Employee

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'personnel_{datetime.now().year}.pdf', 'PDF (*.pdf)')
            if not path: return
            emps = self.session.query(Employee).filter_by(active=True).order_by(Employee.role).all()
            doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
            story = [Paragraph(f'<b>Le Schéma — Liste du Personnel</b>',
                ParagraphStyle('t', fontSize=15, textColor=colors.HexColor('#14B8A6'), alignment=1, spaceAfter=6)),
                Spacer(1,6*mm)]
            data = [['Prénom','Nom','Poste','Téléphone','Email','Embauche','Salaire base']]
            for e in emps:
                data.append([e.first_name or '', e.last_name or '', e.role or '', e.phone or '',
                              e.email or '', str(e.hire_date) if e.hire_date else '', f'{e.base_salary:.0f} MAD'])
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#14B8A6')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
                ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#EAEDF3')),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F9FAFB')]),
                ('PADDING',(0,0),(-1,-1),5),
            ]))
            story.append(t); doc.build(story); self._ok(path)
        except Exception as e: QMessageBox.critical(self, 'Erreur', str(e))

    def _transport_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib import colors; from reportlab.lib.units import mm
            from models.database import Student, Bus

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'transport_{datetime.now().year}.pdf', 'PDF (*.pdf)')
            if not path: return
            students = self.session.query(Student).filter_by(active=True).filter(Student.has_transport == True).all()
            buses    = self.session.query(Bus).filter_by(active=True).all()
            doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
            story = [Paragraph(f'<b>Le Schéma — Rapport Transport</b>',
                ParagraphStyle('t', fontSize=15, textColor=colors.HexColor('#3B82F6'), alignment=1, spaceAfter=4)),
                Paragraph(f'Abonnés: {len(students)} | Bus: {len(buses)}',
                ParagraphStyle('s', fontSize=9, textColor=colors.HexColor('#6B7280'), alignment=1, spaceAfter=6)),
                Spacer(1,6*mm)]
            data = [['Code','Élève','Classe','Parent','Tél.','Frais/mois']]
            for s in students:
                data.append([s.code or '', f'{s.first_name} {s.last_name}', s.class_name or '',
                              s.parent_name or '', s.parent_phone or '',
                              f'{getattr(s,"transport_fee",0):.0f} MAD'])
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#3B82F6')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
                ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#EAEDF3')),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F9FAFB')]),
                ('PADDING',(0,0),(-1,-1),5),
            ]))
            story.append(t); doc.build(story); self._ok(path)
        except Exception as e: QMessageBox.critical(self, 'Erreur', str(e))

    def _payments_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from models.database import Payment, Student, MonthRecord
            from themes.style import SCHOOL_MONTHS

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'paiements_{datetime.now().year}.xlsx', 'Excel (*.xlsx)')
            if not path: return
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Paiements'
            hfill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
            nan_fill  = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
            paid_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
            unpaid_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

            headers = ['Code','Nom','Prénom','Classe'] + SCHOOL_MONTHS + ['Total Payé','Mois Impayés','NAN','Transport','Assurance','Re-inscription']
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = hfill; c.font = Font(color='FFFFFF', bold=True); c.alignment = Alignment(horizontal='center')

            students = self.session.query(Student).filter_by(active=True).order_by(Student.class_name, Student.last_name).all()
            school_year = '2024-25'

            for row_idx, s in enumerate(students, 2):
                records = {r.month_name: r for r in self.session.query(MonthRecord).filter_by(student_id=s.id, school_year=school_year).all()}
                ws.cell(row=row_idx, column=1, value=s.code or '')
                ws.cell(row=row_idx, column=2, value=s.last_name or '')
                ws.cell(row=row_idx, column=3, value=s.first_name or '')
                ws.cell(row=row_idx, column=4, value=s.class_name or '')

                paid_count = 0; unpaid_count = 0; nan_count = 0
                for m_idx, month in enumerate(SCHOOL_MONTHS):
                    r = records.get(month)
                    status = r.status if r else 'unpaid'
                    c = ws.cell(row=row_idx, column=5 + m_idx, value=status.upper())
                    c.alignment = Alignment(horizontal='center')
                    if status == 'paid':    c.fill = paid_fill;   paid_count += 1
                    elif status == 'nan':   c.fill = nan_fill;    nan_count += 1
                    else:                   c.fill = unpaid_fill; unpaid_count += 1

                base_col = 5 + len(SCHOOL_MONTHS)
                transport_str = f'{getattr(s,"transport_fee",0):.0f} MAD' if getattr(s,'has_transport',False) else 'NA'
                from themes.style import REINSCRIPTION_LABELS
                ws.cell(row=row_idx, column=base_col,     value=paid_count)
                ws.cell(row=row_idx, column=base_col+1,   value=unpaid_count)
                ws.cell(row=row_idx, column=base_col+2,   value=nan_count)
                ws.cell(row=row_idx, column=base_col+3,   value=transport_str)
                ws.cell(row=row_idx, column=base_col+4,   value='Payée' if s.insurance_paid else 'Non payée')
                ws.cell(row=row_idx, column=base_col+5,   value=REINSCRIPTION_LABELS.get(getattr(s,'reinscription_status','pending'),'—'))

            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or '')) for c in col)+3, 20)
            wb.save(path); self._ok(path)
        except Exception as e: QMessageBox.critical(self, 'Erreur', str(e))

    def _analytics_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.chart import BarChart, Reference
            from models.database import Payment
            from themes.style import SCHOOL_MONTHS

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'analytiques_{datetime.now().year}.xlsx', 'Excel (*.xlsx)')
            if not path: return
            year = int(self.year_combo.currentText())
            payments = self.session.query(Payment).filter(Payment.year == year).all()
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = f'Analyse {year}'
            hfill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
            ins_fill = PatternFill(start_color='CCFBF1', end_color='CCFBF1', fill_type='solid')
            # Columns: Mois | Mensualités | Transport | Total opérationnel | Assurance (séparé)
            headers = ['Mois', 'Mensualités', 'Transport', 'Total opérationnel', 'Assurance (séparé)']
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                if col == 5:  # insurance column distinct colour
                    c.fill = ins_fill
                    c.font = Font(color='0F766E', bold=True)
                else:
                    c.fill = hfill
                    c.font = Font(color='FFFFFF', bold=True)
                c.alignment = Alignment(horizontal='center')
            for row, m in enumerate(SCHOOL_MONTHS, 2):
                mo  = sum(p.amount for p in payments if p.month==m and p.payment_type=='monthly')
                tr  = sum(p.amount for p in payments if p.month==m and p.payment_type=='transport')
                ins = sum(p.amount for p in payments if p.month==m and p.payment_type=='insurance')
                op  = mo + tr  # operational total (excludes insurance)
                ws.cell(row=row, column=1, value=m)
                ws.cell(row=row, column=2, value=mo)
                ws.cell(row=row, column=3, value=tr)
                ws.cell(row=row, column=4, value=op)
                ins_cell = ws.cell(row=row, column=5, value=ins)
                ins_cell.fill = ins_fill
                ins_cell.font = Font(color='0F766E')
            # Chart only on operational columns (exclude insurance)
            chart = BarChart()
            chart.title = f'Revenus Opérationnels {year} (assurance exclue)'
            chart.y_axis.title = 'MAD'
            chart.add_data(
                Reference(ws, min_col=2, max_col=4, min_row=1, max_row=len(SCHOOL_MONTHS)+1),
                titles_from_data=True
            )
            chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(SCHOOL_MONTHS)+1))
            ws.add_chart(chart, 'G2')
            wb.save(path); self._ok(path)
        except Exception as e: QMessageBox.critical(self, 'Erreur', str(e))

    def _reinscription_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from models.database import Student
            from themes.style import REINSCRIPTION_LABELS

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'reinscription_{datetime.now().year}.xlsx', 'Excel (*.xlsx)')
            if not path: return
            students = self.session.query(Student).filter_by(active=True).order_by(Student.class_name, Student.last_name).all()
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Re-inscription'
            hfill = PatternFill(start_color='EC4899', end_color='EC4899', fill_type='solid')
            for col, h in enumerate(['Code','Nom','Prénom','Classe','Téléphone parent','Statut Re-inscription'], 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = hfill; c.font = Font(color='FFFFFF', bold=True)
            fills = {'yes': 'D1FAE5', 'no': 'FEE2E2', 'pending': 'FEF3C7'}
            for row, s in enumerate(students, 2):
                rs = getattr(s, 'reinscription_status', 'pending') or 'pending'
                ws.cell(row=row, column=1, value=s.code or '')
                ws.cell(row=row, column=2, value=s.last_name or '')
                ws.cell(row=row, column=3, value=s.first_name or '')
                ws.cell(row=row, column=4, value=s.class_name or '')
                ws.cell(row=row, column=5, value=s.parent_phone or '')
                c = ws.cell(row=row, column=6, value=REINSCRIPTION_LABELS.get(rs, rs))
                c.fill = PatternFill(start_color=fills.get(rs,'FFFFFF'), end_color=fills.get(rs,'FFFFFF'), fill_type='solid')
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or '')) for c in col)+3, 28)
            wb.save(path); self._ok(path)
        except Exception as e: QMessageBox.critical(self, 'Erreur', str(e))

    def _nan_report_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from models.database import Student, MonthRecord
            from themes.style import SCHOOL_MONTHS

            path, _ = QFileDialog.getSaveFileName(self, 'Enregistrer', f'rapport_nan_{datetime.now().year}.xlsx', 'Excel (*.xlsx)')
            if not path: return
            students = self.session.query(Student).filter_by(active=True).all()
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Mois NAN'
            nan_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
            hfill    = PatternFill(start_color='854D0E', end_color='854D0E', fill_type='solid')
            for col, h in enumerate(['Code','Nom','Prénom','Classe','Mois NAN','Total mois NAN','Mois payés','Mois impayés'], 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = hfill; c.font = Font(color='FFFFFF', bold=True)
            school_year = '2024-25'
            row_idx = 2
            for s in students:
                records = self.session.query(MonthRecord).filter_by(student_id=s.id, school_year=school_year).all()
                nan_months = [r.month_name for r in records if r.status == 'nan']
                paid_count   = sum(1 for r in records if r.status == 'paid')
                unpaid_count = sum(1 for r in records if r.status == 'unpaid')
                if nan_months:
                    ws.cell(row=row_idx, column=1, value=s.code or '')
                    ws.cell(row=row_idx, column=2, value=s.last_name or '')
                    ws.cell(row=row_idx, column=3, value=s.first_name or '')
                    ws.cell(row=row_idx, column=4, value=s.class_name or '')
                    c = ws.cell(row=row_idx, column=5, value=', '.join(nan_months))
                    c.fill = nan_fill
                    ws.cell(row=row_idx, column=6, value=len(nan_months))
                    ws.cell(row=row_idx, column=7, value=paid_count)
                    ws.cell(row=row_idx, column=8, value=unpaid_count)
                    row_idx += 1
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or '')) for c in col)+3, 35)
            wb.save(path); self._ok(path)
        except Exception as e: QMessageBox.critical(self, 'Erreur', str(e))

    def _backup_db(self):
        import shutil
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        db_path = os.path.join(BASE_DIR, 'database', 'school.db')
        backup_dir = os.path.join(BASE_DIR, 'backups'); os.makedirs(backup_dir, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        dest = os.path.join(backup_dir, f'school_backup_{ts}.db')
        shutil.copy2(db_path, dest)
        self._ok(dest)
        QMessageBox.information(self, 'Sauvegarde ✅', f'Base de données sauvegardée:\n{dest}')
