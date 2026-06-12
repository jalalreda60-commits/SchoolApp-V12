"""
Excel template generator for all modules.
Creates styled, documented templates with sample data.
"""
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from themes.style import CLASSES, SCHOOL_MONTHS, EXPENSE_CATEGORIES

# ── Style helpers ─────────────────────────────────────────────────────
PRIMARY_HEX  = '4F46E5'
SUCCESS_HEX  = '10B981'
WARNING_HEX  = 'F59E0B'
DANGER_HEX   = 'EF4444'
GRAY_HEX     = '6B7280'
LIGHT_HEX    = 'EEF2FF'
NAN_HEX      = 'FEF9C3'

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def _font(bold=False, color='1A1D2E', size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name='Segoe UI')

def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(color='EAEDF3'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _header_cell(ws, row, col, text, bg=PRIMARY_HEX, fg='FFFFFF', width=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill    = _fill(bg)
    cell.font    = _font(bold=True, color=fg, size=10)
    cell.alignment = _align('center')
    cell.border  = _border(bg)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell

def _data_cell(ws, row, col, value, bg='FFFFFF', fg='1A1D2E', bold=False, italic=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill    = _fill(bg)
    cell.font    = _font(bold=bold, color=fg, size=10, italic=italic)
    cell.alignment = _align()
    cell.border  = _border()
    return cell

def _info_sheet(wb, title, columns_info, sample_rows, notes=None):
    """Create an info/legend sheet explaining each column."""
    ws = wb.create_sheet(title='📋 Guide')
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25

    ws.merge_cells('A1:D1')
    cell = ws.cell(row=1, column=1, value=f'📋  Guide du template : {title}')
    cell.fill = _fill(PRIMARY_HEX)
    cell.font = _font(bold=True, color='FFFFFF', size=12)
    cell.alignment = _align('center')

    ws.cell(row=2, column=1, value='Colonne').font = _font(bold=True, color=PRIMARY_HEX)
    ws.cell(row=2, column=2, value='Obligatoire').font = _font(bold=True, color=PRIMARY_HEX)
    ws.cell(row=2, column=3, value='Description').font = _font(bold=True, color=PRIMARY_HEX)
    ws.cell(row=2, column=4, value='Exemple').font = _font(bold=True, color=PRIMARY_HEX)

    for r, (col_name, required, desc, example) in enumerate(columns_info, start=3):
        ws.cell(row=r, column=1, value=col_name).font = _font(bold=True)
        req_cell = ws.cell(row=r, column=2, value='✅ Oui' if required else '⚪ Non')
        req_cell.font = _font(color=SUCCESS_HEX if required else GRAY_HEX)
        ws.cell(row=r, column=3, value=desc)
        ws.cell(row=r, column=4, value=example).font = _font(italic=True, color=GRAY_HEX)
        bg = 'F9FAFB' if r % 2 == 0 else 'FFFFFF'
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = _fill(bg)

    if notes:
        nr = len(columns_info) + 4
        ws.merge_cells(f'A{nr}:D{nr}')
        note_cell = ws.cell(row=nr, column=1, value=f'ℹ️  {notes}')
        note_cell.fill = _fill('FEF9C3')
        note_cell.font = _font(color='854D0E', italic=True)
        note_cell.alignment = _align(wrap=True)


# ═══════════════════════════════════════════════════════════════════════
# 1. STUDENTS TEMPLATE
# ═══════════════════════════════════════════════════════════════════════
def generate_students_template(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '📚 Élèves'
    ws.freeze_panes = 'A2'

    # Fixed headers — matching template_eleves format
    fixed_headers = [
        ('Matricule',        12),
        ('Eleve Nom',        20),
        ('Eleve Prénom',     18),
        ('Mere',             18),
        ('Père',             18),
        ('Date of birth',    14),
        ('City of birth',    16),
        ('Adress',           24),
        ('Père telephone',   16),
        ('Mere telephone',   16),
        ('Classe',           10),
        ('Inscription',      14),
        ('Transport (Y/N)',  14),
        ('Transport',        12),
        ('Mensualité',       14),
        ('Note/Date',        16),
    ]
    month_headers = [(m, 14) for m in SCHOOL_MONTHS]
    all_headers = fixed_headers + month_headers
    n_fixed = len(fixed_headers)

    for col_idx, (name, width) in enumerate(all_headers, start=1):
        bg = PRIMARY_HEX if col_idx <= n_fixed else '312E81'
        _header_cell(ws, 1, col_idx, name, bg=bg, width=width)

    # Legend row (row 2)
    ws.row_dimensions[2].height = 20
    merge_end = get_column_letter(n_fixed)
    ws.merge_cells(f'A2:{merge_end}2')
    legend = ws.cell(row=2, column=1,
        value='⬇ Statuts mois: "Payé" = payé  |  "NAN" = non inscrit ce mois  |  Vide = impayé')
    legend.fill = _fill('FEF9C3')
    legend.font = _font(italic=True, color='854D0E', size=9)
    legend.alignment = _align('center')
    for col in range(n_fixed + 1, len(all_headers) + 1):
        ws.cell(row=2, column=col).fill = _fill('FEF9C3')

    # Data validation — Classe (col 11)
    classes_str = ','.join(CLASSES)
    dv_class = DataValidation(type='list', formula1=f'"{classes_str}"', showDropDown=False)
    dv_class.error = 'Classe invalide. Choisissez dans la liste.'
    dv_class.errorTitle = 'Classe invalide'
    ws.add_data_validation(dv_class)
    dv_class.sqref = 'K3:K2000'

    # Data validation — Transport Y/N (col 13)
    dv_transport_yn = DataValidation(type='list', formula1='"Yes,No"', showDropDown=False)
    dv_transport_yn.error = 'Utilisez: Yes ou No'
    dv_transport_yn.errorTitle = 'Valeur invalide'
    ws.add_data_validation(dv_transport_yn)
    dv_transport_yn.sqref = 'M3:M2000'

    # Data validation — month columns
    dv_months = DataValidation(type='list', formula1='"Payé,NAN,"', showDropDown=False)
    dv_months.error = 'Utilisez: Payé, NAN, ou laissez vide'
    dv_months.errorTitle = 'Valeur invalide'
    ws.add_data_validation(dv_months)
    start_month_col = get_column_letter(n_fixed + 1)
    end_month_col   = get_column_letter(len(all_headers))
    dv_months.sqref = f'{start_month_col}3:{end_month_col}2000'

    # Sample data rows
    SAMPLES = [
        [1, 'ALAOUI',    'Mohamed', 'Fatima ALAOUI',   'Karim ALAOUI',    '15/03/2010', 'Casablanca', '12 Rue Hassan II',    '0612345678', '0698765432', 'CM2',     900,      'Yes', 200,  900,  '01/09/2024',
         'Payé','Payé','Payé','Payé','Payé','Payé','Payé','Payé','Payé',''],
        [2, 'BENALI',    'Fatima',  'Nadia BENALI',    'Omar BENALI',     '22/07/2011', 'Rabat',      '5 Avenue Mohammed V', '0611111111', '0622222222', '6EME',    800,      'No',  0,    800,  '01/09/2024',
         'Payé','Payé','Payé','NAN','NAN','','','','',''],
        [3, 'CHERIF',    'Youssef', 'Zineb CHERIF',    'Hassan CHERIF',   '08/11/2006', 'Marrakech',  '7 Rue Ibn Battuta',   '0633333333', '0644444444', '1BAC SM', 1000,     'Yes', 150,  1000, '01/09/2024',
         'Payé','Payé','Payé','Payé','Payé','Payé','Payé','','',''],
        [4, 'DOUKKALI',  'Sara',    'Amina DOUKKALI',  'Youssef DOUKKALI','14/05/2005', 'Fès',        '3 Bd Al Massira',     '0655555555', '0666666666', '2BAC',    'GRATUIT','No',  0,    0,    '01/09/2024',
         'Payé','Payé','Payé','Payé','Payé','Payé','Payé','Payé','Payé','Payé'],
    ]

    month_bg = {'Payé': 'D1FAE5', 'NAN': 'FEF9C3', '': 'FEE2E2'}
    month_fg = {'Payé': '059669', 'NAN': '854D0E', '': 'DC2626'}

    for r, sample in enumerate(SAMPLES, start=3):
        row_bg = 'F9FAFB' if r % 2 == 0 else 'FFFFFF'
        for c, val in enumerate(sample, start=1):
            if c <= n_fixed:
                _data_cell(ws, r, c, val, bg=row_bg)
            else:
                v = str(val)
                bg = month_bg.get(v, 'FFFFFF')
                fg = month_fg.get(v, '374151')
                _data_cell(ws, r, c, val if val else '',
                           bg=bg if val else row_bg, fg=fg,
                           bold=(val == 'Payé'))

    # Guide sheet
    _info_sheet(wb, 'Élèves', [
        ('Matricule',       False, "Numéro unique de l'élève (auto si vide)",          '1'),
        ('Eleve Nom',       True,  'Nom de famille (MAJUSCULES recommandé)',             'ALAOUI'),
        ('Eleve Prénom',    False, "Prénom de l'élève",                                'Mohamed'),
        ('Mere',            False, 'Nom complet de la mère',                            'Fatima ALAOUI'),
        ('Père',            False, 'Nom complet du père',                               'Karim ALAOUI'),
        ('Date of birth',   False, 'Date de naissance (dd/mm/yyyy)',                    '15/03/2010'),
        ('City of birth',   False, 'Ville de naissance',                                'Casablanca'),
        ('Adress',          False, 'Adresse du domicile',                               '12 Rue Hassan II'),
        ('Père telephone',  False, 'Numéro de téléphone du père',                       '0612345678'),
        ('Mere telephone',  False, 'Numéro de téléphone de la mère',                    '0698765432'),
        ('Classe',          True,  f"Une des classes: {', '.join(CLASSES)}",            'CM2'),
        ('Inscription',     False, "Frais d'inscription (ou GRATUIT)",                 '900'),
        ('Transport (Y/N)', False, 'Transport scolaire: Yes ou No',                     'Yes'),
        ('Transport',       False, 'Montant mensuel transport en MAD (0 si No)',        '200'),
        ('Mensualité',      True,  'Frais mensuel en MAD',                             '900'),
        ('Note/Date',       False, "Date d'inscription (dd/mm/yyyy)",                  '01/09/2024'),
        ('Septembre…Juin',  False, 'Statut mensuel: "Payé", "NAN", ou vide (impayé)',   'Payé'),
    ], [],
    notes='Transport (Y/N): Yes = avec transport, No = sans transport. '
          'Si No, mettre Transport = 0. '
          'Les colonnes des mois acceptent: "Payé", "NAN" (non inscrit), ou vide (impayé).')

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════
# 2. EMPLOYEES TEMPLATE
# ═══════════════════════════════════════════════════════════════════════
def generate_employees_template(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '👥 Personnel'
    ws.freeze_panes = 'A2'

    headers = [
        ('Prénom', 16), ('Nom', 18), ('Poste / Rôle', 18),
        ('Téléphone', 16), ('Email', 24), ('Adresse', 28),
        ('Date embauche', 16), ('Salaire base (MAD)', 20), ('Actif (OUI/NON)', 16),
    ]
    for c, (name, w) in enumerate(headers, 1):
        _header_cell(ws, 1, c, name, bg=SUCCESS_HEX, width=w)

    dv_role = DataValidation(type='list',
        formula1='"Enseignant,Personnel,Chauffeur,Admin,Maintenance"')
    ws.add_data_validation(dv_role); dv_role.sqref = 'C2:C2000'
    dv_active = DataValidation(type='list', formula1='"OUI,NON"')
    ws.add_data_validation(dv_active); dv_active.sqref = 'I2:I2000'

    SAMPLES = [
        ('Karim',  'TAHIRI',   'Enseignant',  '0612345678', 'ktahiri@email.com',   'Casablanca', '01/09/2020', 4500, 'OUI'),
        ('Laila',  'BOUAZZA',  'Enseignant',  '0623456789', 'lbouazza@email.com',  'Rabat',      '15/10/2019', 4200, 'OUI'),
        ('Said',   'MANSOURI', 'Personnel',   '0634567890', '',                    'Casablanca', '01/01/2021', 3000, 'OUI'),
        ('Fatima', 'CHRAIBI',  'Enseignant',  '0645678901', 'fchraibi@email.com',  'Marrakech',  '01/09/2018', 4800, 'OUI'),
        ('Driss',  'KETTANI',  'Chauffeur',   '0656789012', '',                    'Casablanca', '01/03/2022', 2800, 'OUI'),
    ]
    for r, sample in enumerate(SAMPLES, 2):
        bg = 'F9FAFB' if r % 2 == 0 else 'FFFFFF'
        for c, v in enumerate(sample, 1):
            _data_cell(ws, r, c, v, bg=bg)

    _info_sheet(wb, 'Personnel', [
        ('Prénom',         False, 'Prénom de l\'employé',                   'Karim'),
        ('Nom',            True,  'Nom de famille',                         'TAHIRI'),
        ('Poste / Rôle',   True,  'Enseignant | Personnel | Chauffeur | Admin | Maintenance', 'Enseignant'),
        ('Téléphone',      False, 'Numéro de téléphone',                    '0612345678'),
        ('Email',          False, 'Adresse email',                          'k@email.com'),
        ('Adresse',        False, 'Adresse complète',                       'Casablanca'),
        ('Date embauche',  False, 'Format: dd/mm/yyyy',                     '01/09/2020'),
        ('Salaire base',   False, 'Salaire mensuel de base en MAD',         '4500'),
        ('Actif',          False, 'OUI = actif, NON = inactif',             'OUI'),
    ], [])

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════
# 3. EXPENSES TEMPLATE
# ═══════════════════════════════════════════════════════════════════════
def generate_expenses_template(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '💸 Dépenses'
    ws.freeze_panes = 'A2'

    headers = [
        ('Catégorie', 18), ('Type (Fixe/Variable)', 22), ('Description', 32),
        ('Montant (MAD)', 16), ('Date', 14), ('Payé par', 20), ('Notes', 28),
    ]
    for c, (name, w) in enumerate(headers, 1):
        _header_cell(ws, 1, c, name, bg=DANGER_HEX, width=w)

    cats_str = ','.join(EXPENSE_CATEGORIES)
    dv_cat = DataValidation(type='list', formula1=f'"{cats_str}"')
    ws.add_data_validation(dv_cat); dv_cat.sqref = 'A2:A2000'
    dv_type = DataValidation(type='list', formula1='"Fixe,Variable"')
    ws.add_data_validation(dv_type); dv_type.sqref = 'B2:B2000'

    SAMPLES = [
        ('Électricité', 'Fixe',     'Facture ONEE Septembre 2024', 1200, '30/09/2024', 'Direction', ''),
        ('Eau',         'Fixe',     'Facture ONEP Septembre',      350,  '30/09/2024', 'Direction', ''),
        ('Matériel',    'Variable', 'Achat fournitures scolaires', 2500, '05/09/2024', 'Comptable', 'Rentrée scolaire'),
        ('Maintenance', 'Variable', 'Réparation climatisation',    800,  '15/10/2024', 'Technicien', ''),
        ('Communication','Fixe',    'Internet + téléphone',        600,  '01/10/2024', 'Direction', ''),
    ]
    for r, sample in enumerate(SAMPLES, 2):
        bg = 'FEF2F2' if r % 2 == 0 else 'FFFFFF'
        for c, v in enumerate(sample, 1):
            _data_cell(ws, r, c, v, bg=bg)

    _info_sheet(wb, 'Dépenses', [
        ('Catégorie',     True,  f'Une des catégories: {", ".join(EXPENSE_CATEGORIES)}', 'Électricité'),
        ('Type',          True,  '"Fixe" = récurrent mensuel | "Variable" = ponctuel', 'Fixe'),
        ('Description',   False, 'Description détaillée de la dépense', 'Facture ONEE'),
        ('Montant (MAD)', True,  'Montant en dirhams (nombre positif)', '1200'),
        ('Date',          False, 'Date de la dépense (dd/mm/yyyy)',    '30/09/2024'),
        ('Payé par',      False, 'Nom ou poste du responsable',        'Direction'),
        ('Notes',         False, 'Informations supplémentaires',       ''),
    ], [])

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════
# 4. TRANSPORT TEMPLATE
# ═══════════════════════════════════════════════════════════════════════
def generate_transport_template(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '🚌 Transport'
    ws.freeze_panes = 'A2'

    headers = [
        ('Nom du bus / Véhicule', 24), ('Plaque immatriculation', 22),
        ('Capacité (places)', 20), ('Chauffeur (Nom Prénom)', 24), ('Route / Itinéraire', 40),
    ]
    for c, (name, w) in enumerate(headers, 1):
        _header_cell(ws, 1, c, name, bg='0369A1', width=w)

    SAMPLES = [
        ('Bus 1 — Centre', '12345-A-1', 35, 'KETTANI Driss', 'Centre-ville → Hay Mohammadi → École'),
        ('Bus 2 — Nord',   '67890-B-2', 30, 'LAHLOU Hassan', 'Quartier Nord → Avenue Hassan II → École'),
        ('Navette 1',      '11111-C-3', 20, 'BENMOUSSA Ali', 'Zone industrielle → École'),
    ]
    for r, sample in enumerate(SAMPLES, 2):
        bg = 'DBEAFE' if r % 2 == 0 else 'FFFFFF'
        for c, v in enumerate(sample, 1):
            _data_cell(ws, r, c, v, bg=bg)

    _info_sheet(wb, 'Transport', [
        ('Nom du bus',    True,  'Nom ou numéro du véhicule',         'Bus 1 — Centre'),
        ('Plaque',        False, 'Numéro d\'immatriculation',          '12345-A-1'),
        ('Capacité',      False, 'Nombre de places assises',           '35'),
        ('Chauffeur',     False, 'NOM Prénom du chauffeur (doit exister dans Personnel)', 'KETTANI Driss'),
        ('Route',         False, 'Description du trajet et arrêts',    'Centre → École'),
    ], [],
    notes='Le chauffeur doit d\'abord être importé dans le module Personnel avec le rôle "Chauffeur".')

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════
# 5. TIMETABLE TEMPLATE
# ═══════════════════════════════════════════════════════════════════════
def generate_timetable_template(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '📅 Emploi du Temps'
    ws.freeze_panes = 'A2'

    headers = [
        ('Classe', 12), ('Jour', 14), ('Heure début', 14),
        ('Heure fin', 14), ('Matière / Cours', 22),
        ('Enseignant (Nom Prénom)', 26), ('Salle', 14),
    ]
    for c, (name, w) in enumerate(headers, 1):
        _header_cell(ws, 1, c, name, bg='7C3AED', width=w)

    days_str = ','.join(['Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi'])
    dv_day = DataValidation(type='list', formula1=f'"{days_str}"')
    ws.add_data_validation(dv_day); dv_day.sqref = 'B2:B2000'

    classes_str = ','.join(CLASSES)
    dv_cls = DataValidation(type='list', formula1=f'"{classes_str}"')
    ws.add_data_validation(dv_cls); dv_cls.sqref = 'A2:A2000'

    SAMPLES = [
        ('CM2', 'Lundi',    '08:00', '09:00', 'Mathématiques',  'TAHIRI Karim',    'Salle A1'),
        ('CM2', 'Lundi',    '09:00', '10:00', 'Français',       'BOUAZZA Laila',   'Salle A1'),
        ('CM2', 'Lundi',    '10:00', '11:00', 'Sciences',       'TAHIRI Karim',    'Salle A1'),
        ('CM2', 'Mardi',    '08:00', '09:00', 'Histoire-Géo',   'BOUAZZA Laila',   'Salle A1'),
        ('CM2', 'Mardi',    '09:00', '10:00', 'Mathématiques',  'TAHIRI Karim',    'Salle A1'),
        ('6EME','Lundi',    '08:00', '09:00', 'Mathématiques',  'CHRAIBI Fatima',  'Salle B2'),
        ('6EME','Lundi',    '09:00', '10:00', 'Français',       'MANSOURI Said',   'Salle B2'),
        ('1BAC SM','Mercredi','14:00','15:00','Physique-Chimie','TAHIRI Karim',     'Labo'),
    ]
    colors = ['EDE9FE', 'F3F4F6', 'EDE9FE', 'F3F4F6', 'EDE9FE', 'DDD6FE', 'EDE9FE', 'C4B5FD']
    for r, (sample, bg) in enumerate(zip(SAMPLES, colors), 2):
        for c, v in enumerate(sample, 1):
            _data_cell(ws, r, c, v, bg=bg)

    _info_sheet(wb, 'Emploi du Temps', [
        ('Classe',      True,  f'Une des classes: {", ".join(CLASSES[:6])}…', 'CM2'),
        ('Jour',        True,  'Lundi | Mardi | Mercredi | Jeudi | Vendredi | Samedi', 'Lundi'),
        ('Heure début', True,  'Format HH:MM (24h)',                          '08:00'),
        ('Heure fin',   True,  'Format HH:MM (24h)',                          '09:00'),
        ('Matière',     True,  'Nom de la matière enseignée',                 'Mathématiques'),
        ('Enseignant',  False, 'NOM Prénom (doit exister dans Personnel)',    'TAHIRI Karim'),
        ('Salle',       False, 'Nom ou numéro de la salle',                   'Salle A1'),
    ], [],
    notes='L\'enseignant doit d\'abord être importé dans Personnel avec le rôle "Enseignant".')

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════
# 6. PAYMENTS TEMPLATE
# ═══════════════════════════════════════════════════════════════════════
def generate_payments_template(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '💳 Paiements'
    ws.freeze_panes = 'A2'

    headers = [
        ('Code élève', 14), ('Nom élève', 20),
        ('Type paiement', 18), ('Mois', 14), ('Année', 10),
        ('Montant (MAD)', 16), ('Date paiement', 16), ('Notes', 28),
    ]
    for c, (name, w) in enumerate(headers, 1):
        _header_cell(ws, 1, c, name, bg=SUCCESS_HEX, width=w)

    months_str = ','.join(SCHOOL_MONTHS)
    dv_month = DataValidation(type='list', formula1=f'"{months_str}"')
    ws.add_data_validation(dv_month); dv_month.sqref = 'D2:D2000'

    dv_type = DataValidation(type='list',
        formula1='"Mensualite,Assurance,Transport,Inscription"')
    ws.add_data_validation(dv_type); dv_type.sqref = 'C2:C2000'

    SAMPLES = [
        ('STU-0001', 'ALAOUI',    'Mensualite', 'Octobre',   2024, 900,  '05/10/2024', ''),
        ('STU-0001', 'ALAOUI',    'Mensualite', 'Novembre',  2024, 900,  '03/11/2024', ''),
        ('STU-0002', 'BENALI',    'Assurance',  '',          2024, 500,  '01/09/2024', 'Assurance annuelle'),
        ('STU-0003', 'CHERIF',    'Mensualite', 'Septembre', 2024, 1000, '01/09/2024', ''),
        ('STU-0003', 'CHERIF',    'Transport',  'Septembre', 2024, 150,  '01/09/2024', ''),
    ]
    for r, sample in enumerate(SAMPLES, 2):
        bg = 'D1FAE5' if r % 2 == 0 else 'FFFFFF'
        for c, v in enumerate(sample, 1):
            _data_cell(ws, r, c, v, bg=bg)

    _info_sheet(wb, 'Paiements', [
        ('Code élève',    True,  'Code de l\'élève (ex: STU-0001)',          'STU-0001'),
        ('Nom élève',     False, 'Nom de famille (utilisé si code absent)',  'ALAOUI'),
        ('Type paiement', True,  'Mensualite | Assurance | Transport | Inscription', 'Mensualite'),
        ('Mois',          False, f'Un des mois: {", ".join(SCHOOL_MONTHS[:4])}…', 'Octobre'),
        ('Année',         False, 'Année du paiement (ex: 2024 ou 2025)',    '2025'),
        ('Montant (MAD)', True,  'Montant payé en dirhams',                 '900'),
        ('Date paiement', False, 'Date du paiement (dd/mm/yyyy)',           '05/10/2024'),
        ('Notes',         False, 'Informations supplémentaires',            ''),
    ], [],
    notes='Le code élève doit correspondre à un élève existant dans la base de données.')

    wb.save(path)


# ── Dispatcher ────────────────────────────────────────────────────────
TEMPLATE_GENERATORS = {
    'students':   (generate_students_template,   'template_eleves.xlsx',          '📚 Élèves'),
    'employees':  (generate_employees_template,  'template_personnel.xlsx',       '👥 Personnel'),
    'expenses':   (generate_expenses_template,   'template_depenses.xlsx',        '💸 Dépenses'),
    'transport':  (generate_transport_template,  'template_transport.xlsx',       '🚌 Transport'),
    'timetable':  (generate_timetable_template,  'template_emploi_du_temps.xlsx', '📅 Emploi du Temps'),
    'payments':   (generate_payments_template,   'template_paiements.xlsx',       '💳 Paiements'),
}

def generate_template(module: str, save_path: str):
    gen_fn, _, _ = TEMPLATE_GENERATORS[module]
    gen_fn(save_path)
