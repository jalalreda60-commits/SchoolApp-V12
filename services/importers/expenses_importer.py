"""
Expenses Excel importer.
"""
import openpyxl
from models.database import Expense
from services.importers.base_importer import (ImportResult, clean_str,
    clean_float, clean_date)
from themes.style import EXPENSE_CATEGORIES

def _detect_headers(ws):
    row1 = [str(c.value or '').strip().upper() for c in ws[1]]
    m = {}
    for idx, h in enumerate(row1):
        if 'CATEG' in h:                                       m['categorie'] = idx
        elif 'TYPE' in h:                                      m['type'] = idx
        elif 'DESCR' in h or 'LIBELLE' in h or 'OBJET' in h:  m['description'] = idx
        elif 'MONTANT' in h or 'AMOUNT' in h or 'PRIX' in h:  m['montant'] = idx
        elif 'DATE' in h:                                      m['date'] = idx
        elif 'PAYE PAR' in h or 'PAYÉ PAR' in h or 'RESPONSABLE' in h: m['paye_par'] = idx
        elif 'NOTES' in h or 'NOTE' in h or 'COMMENT' in h:   m['notes'] = idx
    return m

def import_expenses(xlsx_path: str, session, mode='skip') -> ImportResult:
    result = ImportResult()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = _detect_headers(ws)

    def get(rv, key, d=None):
        idx = headers.get(key)
        return rv[idx] if idx is not None and idx < len(rv) else d

    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if not any(v for v in row[:5]): continue
        categorie   = clean_str(get(row,'categorie')) or 'Autre'
        exp_type    = clean_str(get(row,'type'), upper=True) or ''
        type_mapped = 'fixed' if 'FIXE' in exp_type or 'FIX' in exp_type else 'variable'
        description = clean_str(get(row,'description'))
        montant     = clean_float(get(row,'montant'))
        exp_date    = clean_date(get(row,'date'))
        paye_par    = clean_str(get(row,'paye_par'))
        notes       = clean_str(get(row,'notes'))

        if montant <= 0:
            result.add_warning(row_num, f"Montant nul ou invalide"); result.skipped += 1; continue

        exp = Expense(
            category=categorie, expense_type=type_mapped,
            description=description, amount=montant,
            date=exp_date, paid_by=paye_par, notes=notes,
        )
        session.add(exp)
        try:
            session.flush(); result.inserted += 1
        except Exception as e:
            session.rollback(); result.add_error(row_num, str(e)); continue

    try:
        session.commit()
    except Exception as e:
        session.rollback(); result.add_error(0, str(e))
    return result
