"""
Timetable (Schedule) Excel importer.
Supports two formats:
  - List format: Class | Day | Start | End | Subject | Teacher | Room
  - Grid format: Time/Day columns matrix
"""
import openpyxl
from models.database import Schedule, Employee
from services.importers.base_importer import ImportResult, clean_str
from themes.style import CLASSES

DAYS_FR = ['Lundi','Mardi','Mercredi','Jeudi','Vendredi','Samedi']

def _detect_headers(ws):
    row1 = [str(c.value or '').strip().upper() for c in ws[1]]
    m = {}
    for idx, h in enumerate(row1):
        if h in ('CLASSE','CLASS'):            m['classe'] = idx
        elif h in ('JOUR','DAY','JOURNÉE'):    m['jour'] = idx
        elif 'DEBUT' in h or 'START' in h or 'DÉBUT' in h: m['debut'] = idx
        elif 'FIN' in h or 'END' in h:        m['fin'] = idx
        elif 'MATIERE' in h or 'MATIÈRE' in h or 'SUBJECT' in h or 'COURS' in h: m['matiere'] = idx
        elif 'ENSEIGNANT' in h or 'PROF' in h or 'TEACHER' in h: m['enseignant'] = idx
        elif 'SALLE' in h or 'ROOM' in h or 'LOCAL' in h: m['salle'] = idx
    return m

def import_timetable(xlsx_path: str, session, mode='skip') -> ImportResult:
    result = ImportResult()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = _detect_headers(ws)

    def get(rv, key, d=None):
        idx = headers.get(key)
        return rv[idx] if idx is not None and idx < len(rv) else d

    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if not any(v for v in row[:5]): continue
        classe   = clean_str(get(row,'classe'), upper=True)
        jour     = clean_str(get(row,'jour'), title=True)
        debut    = clean_str(get(row,'debut'))
        fin      = clean_str(get(row,'fin'))
        matiere  = clean_str(get(row,'matiere'))
        enseignant_name = clean_str(get(row,'enseignant'))
        salle    = clean_str(get(row,'salle'))

        if not classe or not jour or not matiere:
            result.add_error(row_num, "Classe/Jour/Matière manquant"); result.skipped += 1; continue

        if jour not in DAYS_FR:
            result.add_warning(row_num, f"Jour inconnu: '{jour}'")

        # Find teacher
        teacher_id = None
        if enseignant_name:
            parts = enseignant_name.split()
            q = session.query(Employee).filter(Employee.role.in_(['teacher']))
            if len(parts) >= 2:
                q = q.filter(Employee.last_name.ilike(f'%{parts[-1]}%'))
            elif len(parts) == 1:
                q = q.filter((Employee.last_name.ilike(f'%{parts[0]}%')) |
                              (Employee.first_name.ilike(f'%{parts[0]}%')))
            teacher = q.first()
            if teacher: teacher_id = teacher.id

        existing = session.query(Schedule).filter_by(
            class_name=classe, day=jour, time_start=debut, subject=matiere
        ).first()

        if existing and mode == 'skip':
            result.skipped += 1; continue
        elif existing and mode == 'update':
            existing.time_end=fin or existing.time_end
            existing.teacher_id=teacher_id or existing.teacher_id
            existing.room=salle or existing.room
            result.updated += 1
        else:
            sch = Schedule(class_name=classe, day=jour, time_start=debut,
                time_end=fin, subject=matiere, teacher_id=teacher_id, room=salle)
            session.add(sch)
            try:
                session.flush(); result.inserted += 1
            except Exception as e:
                session.rollback(); result.add_error(row_num, str(e)); continue

    try:
        session.commit()
    except Exception as e:
        session.rollback(); result.add_error(0, str(e))
    return result
