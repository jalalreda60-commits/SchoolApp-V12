"""
Payments Excel importer.
Adds payment records for existing students.
"""
import openpyxl
from datetime import datetime
from models.database import Student, Payment, MonthRecord
from services.importers.base_importer import (ImportResult, clean_str, clean_float, clean_date)
from themes.style import SCHOOL_MONTHS

MONTH_MAP = {m.upper()[:4]: m for m in SCHOOL_MONTHS}
MONTH_CAL = {'Septembre':9,'Octobre':10,'Novembre':11,'Décembre':12,
             'Janvier':1,'Février':2,'Mars':3,'Avril':4,'Mai':5,'Juin':6}

def _detect_headers(ws):
    row1 = [str(c.value or '').strip().upper() for c in ws[1]]
    m = {}
    for idx, h in enumerate(row1):
        if 'CODE' in h or 'MATRICULE' in h:           m['code'] = idx
        elif 'NOM' in h and 'FAMILLE' not in h:       m['nom'] = idx
        elif 'TYPE' in h or 'PAIEMENT' in h:          m['type'] = idx
        elif 'MOIS' in h or 'MONTH' in h or 'PERIODE' in h: m['mois'] = idx
        elif 'ANNEE' in h or 'ANNÉE' in h or 'YEAR' in h: m['annee'] = idx
        elif 'MONTANT' in h or 'AMOUNT' in h:         m['montant'] = idx
        elif 'DATE' in h:                              m['date'] = idx
        elif 'NOTES' in h or 'NOTE' in h:             m['notes'] = idx
    return m

def _find_student(session, code_raw, nom_raw):
    if code_raw:
        code_str = str(code_raw).strip()
        s = session.query(Student).filter_by(code=code_str).first()
        if s: return s
        # Try as matricule
        mat = code_str.zfill(4)
        s = session.query(Student).filter(Student.code.endswith(mat)).first()
        if s: return s
    if nom_raw:
        nom = str(nom_raw).strip().upper()
        s = session.query(Student).filter_by(last_name=nom, active=True).first()
        if s: return s
    return None

def import_payments(xlsx_path: str, session, mode='skip') -> ImportResult:
    result = ImportResult()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = _detect_headers(ws)
    pay_count = session.query(Payment).count()

    def get(rv, key, d=None):
        idx = headers.get(key)
        return rv[idx] if idx is not None and idx < len(rv) else d

    TYPE_MAP = {
        'MENSUALITE':'monthly','MENSUALITÉ':'monthly','MONTHLY':'monthly',
        'ASSURANCE':'insurance','INSURANCE':'insurance',
        'TRANSPORT':'transport',
        'INSCRIPTION':'inscription',
    }

    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if not any(v for v in row[:5]): continue
        code    = get(row,'code')
        nom     = get(row,'nom')
        type_raw= clean_str(get(row,'type'), upper=True) or 'MONTHLY'
        mois    = clean_str(get(row,'mois'), title=True)
        annee   = int(clean_float(get(row,'annee')) or datetime.now().year)
        montant = clean_float(get(row,'montant'))
        pay_date= clean_date(get(row,'date')) or datetime.now().date()
        notes   = clean_str(get(row,'notes'))

        pay_type = TYPE_MAP.get(type_raw, 'monthly')

        # Resolve month
        month_name = None
        if mois:
            mois_up = mois.upper()[:4]
            month_name = MONTH_MAP.get(mois_up, mois)

        student = _find_student(session, code, nom)
        if not student:
            result.add_error(row_num, f"Élève introuvable: code={code} nom={nom}")
            result.skipped += 1; continue

        if montant <= 0:
            result.add_warning(row_num, f"Montant nul pour {student.code}"); result.skipped += 1; continue

        # Duplicate check
        if pay_type == 'monthly' and month_name:
            existing = session.query(Payment).filter_by(
                student_id=student.id, payment_type='monthly', month=month_name, year=annee
            ).first()
            if existing and mode == 'skip':
                result.skipped += 1; continue

        pay_count += 1
        rec_num = f'REC-IMP-{pay_count:06d}'
        pay = Payment(
            student_id=student.id, payment_type=pay_type,
            amount=montant, month=month_name, year=annee,
            payment_date=datetime.combine(pay_date, datetime.min.time()),
            receipt_number=rec_num, notes=notes,
        )
        session.add(pay)

        # Sync MonthRecord
        if pay_type == 'monthly' and month_name:
            mr = session.query(MonthRecord).filter_by(
                student_id=student.id, month_name=month_name
            ).first()
            if mr: mr.status = 'paid'; mr.amount = montant
        elif pay_type == 'insurance':
            student.insurance_paid = True

        try:
            session.flush(); result.inserted += 1
        except Exception as e:
            session.rollback(); result.add_error(row_num, str(e)); continue

    try:
        session.commit()
    except Exception as e:
        session.rollback(); result.add_error(0, str(e))
    return result
