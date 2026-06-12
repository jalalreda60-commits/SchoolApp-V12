"""
Employees Excel importer.
"""
import openpyxl
from models.database import Employee
from services.importers.base_importer import (ImportResult, clean_str,
    clean_float, clean_date, clean_bool)

ROLE_MAP = {
    'ENSEIGNANT':'teacher','TEACHER':'teacher','PROF':'teacher','PROFESSEUR':'teacher',
    'PERSONNEL':'staff','STAFF':'staff','ADMIN':'admin','ADMINISTRATION':'admin',
    'CHAUFFEUR':'driver','DRIVER':'driver','MAINTENANCE':'maintenance',
    'TECHNICIEN':'maintenance','SECRETAIRE':'staff','COMPTABLE':'staff',
}

def _detect_headers(ws):
    row1 = [str(c.value or '').strip().upper() for c in ws[1]]
    m = {}
    for idx, h in enumerate(row1):
        if 'PRENOM' in h or 'PRÉNOM' in h or h == 'FIRSTNAME': m['prenom'] = idx
        elif h in ('NOM','NOM DE FAMILLE','LASTNAME','NAME'):   m['nom'] = idx
        elif 'POSTE' in h or 'ROLE' in h or 'FONCTION' in h:   m['role'] = idx
        elif 'TEL' in h or 'TÉL' in h or 'PHONE' in h or 'PORTABLE' in h: m['phone'] = idx
        elif 'EMAIL' in h or 'MAIL' in h:                       m['email'] = idx
        elif 'ADRESSE' in h or 'ADDRESS' in h:                 m['adresse'] = idx
        elif 'EMBAUCHE' in h or 'HIRE' in h or 'RECRUTEMENT' in h: m['hire_date'] = idx
        elif 'SALAIRE' in h or 'SALARY' in h:                   m['salaire'] = idx
        elif 'ACTIF' in h or 'ACTIVE' in h or 'STATUT' in h:    m['actif'] = idx
    return m

def import_employees(xlsx_path: str, session, mode='skip') -> ImportResult:
    result = ImportResult()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = _detect_headers(ws)

    def get(rv, key, d=None):
        idx = headers.get(key)
        return rv[idx] if idx is not None and idx < len(rv) else d

    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if not any(v for v in row[:6]): continue
        first_name = clean_str(get(row,'prenom'), title=True)
        last_name  = clean_str(get(row,'nom'),    upper=True)
        role_raw   = clean_str(get(row,'role'), upper=True) or ''
        role       = ROLE_MAP.get(role_raw, 'staff')
        phone      = clean_str(get(row,'phone'))
        email      = clean_str(get(row,'email'))
        adresse    = clean_str(get(row,'adresse'))
        hire_date  = clean_date(get(row,'hire_date'))
        salaire    = clean_float(get(row,'salaire'))
        actif      = clean_bool(get(row,'actif')) if get(row,'actif') is not None else True

        if not last_name:
            result.add_error(row_num, "Nom manquant"); result.skipped += 1; continue

        existing = session.query(Employee).filter_by(
            last_name=last_name, first_name=first_name or '', role=role
        ).first()

        if existing and mode == 'skip':
            result.skipped += 1; continue
        elif existing and mode == 'update':
            existing.phone=phone or existing.phone; existing.email=email or existing.email
            existing.base_salary=salaire or existing.base_salary; result.updated += 1
        else:
            emp = Employee(
                first_name=first_name or '', last_name=last_name, role=role,
                phone=phone, email=email, address=adresse,
                hire_date=hire_date, base_salary=salaire, active=actif,
            )
            session.add(emp)
            try:
                session.flush(); result.inserted += 1
            except Exception as e:
                session.rollback(); result.add_error(row_num, str(e)); continue

    try:
        session.commit()
    except Exception as e:
        session.rollback(); result.add_error(0, str(e))
    return result
