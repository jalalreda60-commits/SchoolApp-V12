"""
students_importer.py — SGS v4
Excel import rules:
  Payé  → MonthRecord(status='paid')  + Payment(monthly) + Payment(transport if applicable)
  NAN   → MonthRecord(status='nan')
  Empty → MonthRecord(status='unpaid')

Financial history is automatically reconstructed from 'Payé' columns so
that the dashboard can calculate revenue from imported data immediately.
"""
import openpyxl
from datetime import date, datetime
from models.database import Student, MonthRecord, Payment, Setting, SCHOOL_MONTHS
from services.importers.base_importer import (ImportResult, clean_str,
    clean_float, clean_date)
from themes.style import CLASSES

CLASS_ALIASES = {
    'CE6': '6EME', '6ÈME': '6EME', '6EME': '6EME',
    '1BACSM': '1BAC SM', '1BAC-SM': '1BAC SM',
}

REQUIRED_HEADERS = {'nom'}

# School month index → (calendar_month, calendar_year_offset)
# offset 0 = same year as school year end, -1 = previous year (Sep-Dec)
_MONTH_CAL = [
    (9, -1), (10, -1), (11, -1), (12, -1),   # Sep Oct Nov Dec → prev year
    (1,  0), (2,  0),  (3,  0),  (4,  0),    # Jan Feb Mar Apr → current year
    (5,  0), (6,  0),                          # Mai Jun → current year
]


def _detect_headers(ws):
    row1 = [str(c.value or '').strip().upper() for c in ws[1]]
    mapping = {}
    for idx, h in enumerate(row1):
        if 'MATRICULE' in h or h in ('MAT', 'N°', 'NUM'):
            mapping['matricule'] = idx
        # Nom — support "Eleve Nom", "Nom", "Lastname"
        elif h in ('ELEVE NOM', 'NOM', 'NOM DE FAMILLE', 'LASTNAME', 'NAME', 'ÉLÈVE NOM'):
            mapping['nom'] = idx
        # Prénom — support "Eleve Prénom", "Prénom", "Firstname"
        elif h in ('ELEVE PRÉNOM', 'ELEVE PRENOM', 'PRENOM', 'PRÉNOM', 'FIRSTNAME'):
            mapping['prenom'] = idx
        # Parents
        elif h in ('MERE', 'MÈRE', 'NOM MERE', 'NOM MÈRE', 'MOTHER'):
            mapping['mere'] = idx
        elif h in ('PÈRE', 'PERE', 'NOM PÈRE', 'NOM PERE', 'FATHER'):
            mapping['pere'] = idx
        elif h in ('DATE OF BIRTH', 'DATE NAISSANCE', 'NAISSANCE', 'BIRTH', 'DOB'):
            mapping['naissance'] = idx
        elif h in ('CITY OF BIRTH', 'VILLE NAISSANCE', 'CITY', 'VILLE'):
            mapping['ville_naissance'] = idx
        elif h in ('ADRESS', 'ADRESSE', 'ADDRESS'):
            mapping['adresse'] = idx
        elif h in ('PÈRE TELEPHONE', 'PERE TELEPHONE', 'PÈRE TEL', 'PERE TEL',
                   'FATHER PHONE', 'FATHER TEL', 'TEL PÈRE', 'TEL PERE'):
            mapping['pere_tel'] = idx
        elif h in ('MERE TELEPHONE', 'MÈRE TELEPHONE', 'MÈRE TEL', 'MERE TEL',
                   'MOTHER PHONE', 'MOTHER TEL', 'TEL MERE', 'TEL MÈRE'):
            mapping['mere_tel'] = idx
        # Legacy single parent phone (fallback if new columns absent)
        elif ('TÉL' in h or 'TEL' in h or 'PHONE' in h or 'PORTABLE' in h) and 'PARENT' in h:
            mapping['parent_phone'] = idx
        elif h == 'CLASSE' or h == 'CLASS':
            mapping['classe'] = idx
        elif 'INSCRIPT' in h:
            mapping['inscription'] = idx
        elif h == 'TRANSPORT (Y/N)' or h == 'TRANSPORT Y/N':
            mapping['transport_yn'] = idx
        elif h == 'TRANSPORT':
            mapping['transport'] = idx
        elif 'MENSUAL' in h or h == 'FRAIS' or 'FRAIS MENS' in h or 'MENSUALITÉ' in h:
            mapping['mensualite'] = idx
        elif 'ASSURANCE' in h:
            mapping['assurance'] = idx
        elif h in ('NOTE', 'DATE', 'NOTES', 'DATE INSCR', 'NOTE/DATE'):
            mapping['note'] = idx
        elif 'PARENT' in h and 'NOM' in h:
            mapping['parent_nom'] = idx
        elif ('TÉL' in h or 'TEL' in h or 'PHONE' in h or 'PORTABLE' in h):
            if 'parent_phone' not in mapping and 'pere_tel' not in mapping and 'mere_tel' not in mapping:
                mapping['parent_phone'] = idx
        elif 'URGENCE' in h:
            mapping['urgence'] = idx
        elif 'GENRE' in h or 'SEXE' in h or 'GENDER' in h:
            mapping['genre'] = idx
        elif 'REINSCR' in h or 'RE-INSCR' in h:
            mapping['reinscription'] = idx
        else:
            for m_idx, month in enumerate(SCHOOL_MONTHS):
                m_up = month.upper()
                if h == m_up or h.startswith(m_up[:4]):
                    mapping[f'month_{m_idx}'] = idx
                    break
    return mapping

def _clean_class(raw) -> str:
    if not raw:
        return None
    s = str(raw).strip().upper().replace(' ', '')
    mapped = CLASS_ALIASES.get(s, CLASS_ALIASES.get(s.replace('-', ''), s))
    for c in CLASSES:
        if c.upper().replace(' ', '') == mapped:
            return c
    return str(raw).strip()


def _parse_month_status(cell, col_value, is_nan_color):
    """
    Returns: 'paid' | 'nan' | 'unpaid'
    Rules:
      Payé        → paid
      NAN / color → nan
      Empty       → unpaid
    """
    val = str(col_value or '').strip().upper()
    if is_nan_color or val == 'NAN':
        return 'nan'
    if val in ('PAYÉ', 'PAYE', 'PAYEE', 'PAYÉE', 'P', 'OUI', 'YES', '1'):
        return 'paid'
    return 'unpaid'


def _get_school_year(session):
    s = session.query(Setting).filter_by(key='school_year').first()
    return s.value if s else '2024-25'


def _school_year_to_end_year(school_year: str) -> int:
    """'2024-25' → 2025"""
    try:
        return int(school_year.split('-')[0]) + 1
    except Exception:
        return datetime.now().year


def _payment_exists(session, student_id, month_name, payment_type):
    """Duplicate check: student + month + payment_type."""
    return session.query(Payment).filter_by(
        student_id=student_id,
        payment_type=payment_type,
        month=month_name,
    ).first() is not None


def import_students(xlsx_path: str, session, mode='skip') -> ImportResult:
    """
    mode: 'skip'   = keep existing records unchanged
          'update' = update financial fields on existing students
    """
    result = ImportResult()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = _detect_headers(ws)
    school_year = _get_school_year(session)
    end_year    = _school_year_to_end_year(school_year)   # e.g. 2025 for '2024-25'

    def get(row_vals, key, default=None):
        idx = headers.get(key)
        return row_vals[idx] if idx is not None and idx < len(row_vals) else default

    def get_cell(row_cells, key):
        idx = headers.get(key)
        return row_cells[idx] if idx is not None and idx < len(row_cells) else None

    pay_count = session.query(Payment).count()

    for row_num, row_cells in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        row_vals = [c.value for c in row_cells]

        # Skip blank/legend rows
        if not any(v for v in row_vals[:8]):
            continue
        first_str = str(row_vals[0] or '').strip()
        if first_str.startswith('⬇') or 'Statuts' in first_str:
            continue

        # ── Parse student fields ──────────────────────────────────────────────
        mat_raw       = get(row_vals, 'matricule')
        last_name     = clean_str(get(row_vals, 'nom'), upper=True)
        first_name    = clean_str(get(row_vals, 'prenom'), title=True)
        classe_raw    = get(row_vals, 'classe')
        classe        = _clean_class(classe_raw)
        mensualite    = clean_float(get(row_vals, 'mensualite')) or 0.0

        # Transport: new template has "Transport (Y/N)" + "Transport" (amount) columns
        transport_yn_raw  = get(row_vals, 'transport_yn')   # 'Yes' / 'No'
        transport_raw     = get(row_vals, 'transport')       # numeric amount
        transport_amt     = clean_float(transport_raw) or 0.0
        if transport_yn_raw is not None:
            has_transport = str(transport_yn_raw or '').upper() in ('YES', 'OUI', 'O', 'Y')
        else:
            # legacy single-column: positive amount → has transport
            has_transport = (transport_amt > 0
                             or str(transport_raw or '').upper() in ('OUI', 'YES', 'O'))

        assurance     = clean_float(get(row_vals, 'assurance'), default=500.0)

        # Parents (new template)
        mere_nom      = clean_str(get(row_vals, 'mere'))
        pere_nom      = clean_str(get(row_vals, 'pere'))
        pere_tel      = clean_str(get(row_vals, 'pere_tel'))
        mere_tel      = clean_str(get(row_vals, 'mere_tel'))
        # Legacy fallback
        parent_nom    = clean_str(get(row_vals, 'parent_nom')) or pere_nom or mere_nom
        parent_tel    = clean_str(get(row_vals, 'parent_phone')) or pere_tel or mere_tel

        urgence       = clean_str(get(row_vals, 'urgence'))
        genre         = clean_str(get(row_vals, 'genre'))
        naissance     = clean_date(get(row_vals, 'naissance'))
        ville_naissance = clean_str(get(row_vals, 'ville_naissance'))
        adresse       = clean_str(get(row_vals, 'adresse'))
        reg_date      = clean_date(get(row_vals, 'note'))
        reinscr       = clean_str(get(row_vals, 'reinscription'))
        reinscr_status = {'oui': 'yes', 'non': 'no', 'yes': 'yes', 'no': 'no'}.get(
            str(reinscr or '').lower(), 'pending'
        )

        if not last_name:
            result.add_error(row_num, "Nom manquant — ligne ignorée")
            result.skipped += 1
            continue

        if classe and classe not in CLASSES:
            result.add_warning(row_num, f"Classe inconnue: '{classe}'")

        # ── Build code / find existing student ───────────────────────────────
        mat_str = (str(int(float(mat_raw)))
                   if mat_raw and str(mat_raw).replace('.', '').isdigit() else '')
        code = f'STU-{mat_str.zfill(4)}' if mat_str else None

        existing = None
        if code:
            existing = session.query(Student).filter_by(code=code).first()
        if not existing and last_name:
            existing = session.query(Student).filter_by(
                last_name=last_name, class_name=classe, active=True
            ).first()

        if existing and mode == 'skip':
            result.skipped += 1
            student = existing
        elif existing and mode == 'update':
            existing.first_name       = first_name or existing.first_name
            existing.monthly_fee      = mensualite or existing.monthly_fee
            existing.has_transport    = has_transport
            existing.transport_fee    = transport_amt
            existing.insurance_amount = assurance
            existing.mother_name      = mere_nom or existing.mother_name
            existing.father_name      = pere_nom or existing.father_name
            existing.mother_phone     = mere_tel or existing.mother_phone
            existing.father_phone     = pere_tel or existing.father_phone
            existing.parent_name      = parent_nom or existing.parent_name
            existing.parent_phone     = parent_tel or existing.parent_phone
            existing.city_of_birth    = ville_naissance or existing.city_of_birth
            result.updated += 1
            student = existing
        else:
            student = Student(
                code=code or f'STU-IMP-{row_num:04d}',
                first_name=first_name or '',
                last_name=last_name,
                gender=genre,
                birth_date=naissance,
                city_of_birth=ville_naissance,
                class_name=classe,
                address=adresse,
                mother_name=mere_nom,
                father_name=pere_nom,
                mother_phone=mere_tel,
                father_phone=pere_tel,
                parent_name=parent_nom,
                parent_phone=parent_tel,
                emergency_phone=urgence,
                monthly_fee=mensualite,
                has_transport=has_transport,
                transport_fee=transport_amt,
                insurance_amount=assurance,
                insurance_paid=False,
                reinscription_status=reinscr_status,
                registration_date=reg_date or date(2024, 9, 1),
                notes=f'MAT:{mat_str}' if mat_str else None,
                active=True,
            )
            session.add(student)
            try:
                session.flush()
                result.inserted += 1
            except Exception as e:
                session.rollback()
                result.add_error(row_num, str(e))
                continue

        # ── Month records & historical payment reconstruction ─────────────────
        for m_idx, month_name in enumerate(SCHOOL_MONTHS):
            cell     = get_cell(row_cells, f'month_{m_idx}')
            cell_val = cell.value if cell else None

            # Detect NAN highlight colour (yellow = FFFFFF00)
            try:
                is_nan_color = cell and cell.fill.fgColor.rgb == 'FFFFFF00'
            except Exception:
                is_nan_color = False

            status = _parse_month_status(cell, cell_val, is_nan_color)

            # Derive calendar year for this school month
            cal_month, year_offset = _MONTH_CAL[m_idx]
            pay_year = end_year + year_offset   # Sep-Dec → end_year-1, Jan-Jun → end_year

            # ── MonthRecord ──────────────────────────────────────────────────
            month_amount = mensualite if status != 'nan' else 0.0
            mr = session.query(MonthRecord).filter_by(
                student_id=student.id,
                month_name=month_name,
                school_year=school_year,
            ).first()

            if mr:
                mr.status = status
                mr.amount = month_amount
            else:
                session.add(MonthRecord(
                    student_id=student.id,
                    month_name=month_name,
                    school_year=school_year,
                    status=status,
                    amount=month_amount,
                ))

            # ── Historical Payment records — only for 'paid' months ───────────
            # Creates Payment(monthly) and optionally Payment(transport)
            # Duplicate prevention: student + month + payment_type
            if status == 'paid' and mensualite > 0:

                # 1) Monthly payment
                if not _payment_exists(session, student.id, month_name, 'monthly'):
                    pay_count += 1
                    session.add(Payment(
                        student_id=student.id,
                        payment_type='monthly',
                        amount=mensualite,
                        month=month_name,
                        year=pay_year,
                        school_year=school_year,
                        payment_date=datetime(pay_year, cal_month, 1),
                        receipt_number=f'REC-IMP-{pay_count:06d}',
                    ))

                # 2) Transport payment (only if student has transport)
                if has_transport and transport_amt > 0:
                    if not _payment_exists(session, student.id, month_name, 'transport'):
                        pay_count += 1
                        session.add(Payment(
                            student_id=student.id,
                            payment_type='transport',
                            amount=transport_amt,
                            month=month_name,
                            year=pay_year,
                            school_year=school_year,
                            payment_date=datetime(pay_year, cal_month, 1),
                            receipt_number=f'REC-IMP-{pay_count:06d}',
                        ))

    try:
        session.commit()
    except Exception as e:
        session.rollback()
        result.add_error(0, f"Commit error: {e}")

    return result
