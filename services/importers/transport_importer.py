"""
Transport (Buses) Excel importer.
"""
import openpyxl
from models.database import Bus, Employee
from services.importers.base_importer import ImportResult, clean_str, clean_int

def _detect_headers(ws):
    row1 = [str(c.value or '').strip().upper() for c in ws[1]]
    m = {}
    for idx, h in enumerate(row1):
        if 'NOM' in h or 'BUS' in h or 'VEHICULE' in h or 'VÉHICULE' in h: m['nom'] = idx
        elif 'PLAQUE' in h or 'IMMAT' in h or 'PLATE' in h:  m['plaque'] = idx
        elif 'CAPAC' in h or 'PLACES' in h or 'SEATS' in h:  m['capacite'] = idx
        elif 'CHAUFFEUR' in h or 'DRIVER' in h or 'CONDUCTEUR' in h: m['chauffeur'] = idx
        elif 'ROUTE' in h or 'TRAJET' in h or 'ITIN' in h:   m['route'] = idx
    return m

def import_buses(xlsx_path: str, session, mode='skip') -> ImportResult:
    result = ImportResult()
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = _detect_headers(ws)

    def get(rv, key, d=None):
        idx = headers.get(key)
        return rv[idx] if idx is not None and idx < len(rv) else d

    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if not any(v for v in row[:4]): continue
        nom      = clean_str(get(row,'nom'))
        plaque   = clean_str(get(row,'plaque'), upper=True)
        capacite = clean_int(get(row,'capacite')) or 30
        chauffeur_name = clean_str(get(row,'chauffeur'))
        route    = clean_str(get(row,'route'))

        if not nom:
            result.add_error(row_num, "Nom bus manquant"); result.skipped += 1; continue

        existing = session.query(Bus).filter_by(name=nom).first()
        if existing and mode == 'skip':
            result.skipped += 1; continue

        # Find driver
        driver_id = None
        if chauffeur_name:
            parts = chauffeur_name.split()
            q = session.query(Employee).filter_by(role='driver')
            if len(parts) >= 2:
                q = q.filter(Employee.last_name.ilike(f'%{parts[-1]}%'))
            driver = q.first()
            if driver: driver_id = driver.id

        if existing and mode == 'update':
            existing.plate=plaque or existing.plate; existing.capacity=capacite
            existing.driver_id=driver_id or existing.driver_id; existing.route=route or existing.route
            result.updated += 1
        else:
            bus = Bus(name=nom, plate=plaque, capacity=capacite, driver_id=driver_id, route=route, active=True)
            session.add(bus)
            try:
                session.flush(); result.inserted += 1
            except Exception as e:
                session.rollback(); result.add_error(row_num, str(e)); continue

    try:
        session.commit()
    except Exception as e:
        session.rollback(); result.add_error(0, str(e))
    return result
