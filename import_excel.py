#!/usr/bin/env python3
"""
Le Schéma SGS v3 — Excel Data Importer
Imports students, payment statuses, and financial data from sch.xlsx
"""
import sys
import os
import hashlib
from datetime import datetime, date

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

# ── Column mapping ─────────────────────────────────────────────────────
COL = {
    'matricule':   0,   # A
    'nom':         1,   # B  (last name)
    'prenom':      2,   # C  (first name)
    'classe':      3,   # D
    'inscription': 4,   # E  (registration fee)
    'transport':   5,   # F
    'mensualite':  6,   # G  (monthly fee)
    'total':       7,   # H
    'note':        8,   # I  (registration date or note)
    # Months: cols J(9)…S(18) = Sep…Jun — 10 months
}

MONTH_COLS = list(range(9, 19))  # indices 9..18

SCHOOL_MONTHS_10 = [
    'Septembre','Octobre','Novembre','Décembre',
    'Janvier','Février','Mars','Avril','Mai','Juin'
]

# Class name fixes
CLASS_MAP = {
    'CE6': '6EME',   # CE6 appears to be 6ème
    '':    None,
}

VALID_CLASSES = {
    'PS','MS','GS','CP','CE1','CE2','CM1','CM2',
    '6EME','1AC','2AC','3AC','TC','1BAC','1BAC SM','2BAC'
}


def clean_class(raw):
    raw = str(raw).strip() if raw else ''
    mapped = CLASS_MAP.get(raw, raw)
    if mapped in VALID_CLASSES:
        return mapped
    # fuzzy fix
    upper = raw.upper()
    if upper in VALID_CLASSES:
        return upper
    if raw == 'CE6':
        return '6EME'
    return raw if raw else None


def clean_amount(val):
    if val is None:
        return 0.0
    s = str(val).strip().upper()
    if not s or s in ('NONE', 'GRATUIT', '0', ''):
        return 0.0
    # Strip text suffixes like "500 AV", "1000 AV", "3000DH AV"
    import re
    m = re.search(r'[\d]+(?:\.\d+)?', s)
    if m:
        return float(m.group())
    return 0.0


def clean_transport(val):
    if val is None:
        return 0.0, False
    s = str(val).strip().upper()
    if s in ('0', '', 'NONE', 'OUI'):
        # OUI without amount means has transport but fee unknown
        if s == 'OUI':
            return 0.0, True
        return 0.0, False
    try:
        amt = float(s)
        return (amt, amt > 0)
    except ValueError:
        return 0.0, False


def clean_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    if not s or s in ('None', ''):
        return None
    # Try various formats
    for fmt in ('%d/%m%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_month_status(cell_val, is_yellow_nan):
    """
    Returns: 'paid' | 'unpaid' | 'nan'
    Logic:
      - Yellow cell (FFFFFF00) OR value=='NAN' → 'nan'
      - value=='Payé'                           → 'paid'
      - value==None/''                          → 'unpaid'
    """
    val = str(cell_val).strip().upper() if cell_val is not None else ''
    if is_yellow_nan or val == 'NAN':
        return 'nan'
    if val in ('PAYÉ', 'PAYE', 'PAYEE', 'PAYÉE'):
        return 'paid'
    return 'unpaid'


def run_import(xlsx_path, clear_existing=False, school_year='2024-25'):
    from models.database import (init_db, migrate_db, get_session,
        Student, MonthRecord, Payment, Receipt, Setting, User, Base, engine)

    print("=" * 60)
    print("  Le Schéma SGS v3 — Excel Importer")
    print("=" * 60)

    # ── Step 1: Init DB ───────────────────────────────────────────
    print("\n[1/6] Initialising database...")
    init_db()
    migrate_db()
    session = get_session()
    print("  ✅ Database ready")

    # ── Step 2: Load Excel ────────────────────────────────────────
    print(f"\n[2/6] Loading Excel: {xlsx_path}")
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    total_rows = ws.max_row - 1
    print(f"  Found {total_rows} data rows (excluding header)")

    # ── Step 3: Optionally clear existing students ────────────────
    if clear_existing:
        print("\n[3/6] Clearing existing student data...")
        session.query(MonthRecord).delete()
        session.query(Receipt).delete()
        session.query(Payment).delete()
        session.query(Student).delete()
        session.commit()
        print("  ✅ Cleared")
    else:
        print("\n[3/6] Keeping existing data (will skip duplicates)")

    # ── Step 4: Import students ───────────────────────────────────
    print(f"\n[4/6] Importing students (school year: {school_year})...")

    stats = {
        'inserted': 0, 'skipped': 0, 'errors': 0,
        'paid_months': 0, 'nan_months': 0, 'unpaid_months': 0,
        'payments_created': 0,
    }

    existing_codes = {s.code for s in session.query(Student).all()}
    existing_mats  = {s.notes: s for s in session.query(Student).all() if s.notes}  # use notes to store matricule

    for excel_row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        mat_cell = excel_row[0]
        if not mat_cell.value:
            continue  # skip blank rows

        # ── Parse row ─────────────────────────────────────────────
        row = [cell.value for cell in excel_row]
        matricule = str(int(row[COL['matricule']])) if isinstance(row[COL['matricule']], float) else str(row[COL['matricule']] or '')
        last_name  = str(row[COL['nom']]    or '').strip().upper()
        first_name = str(row[COL['prenom']] or '').strip().title()
        classe     = clean_class(row[COL['classe']])
        mensualite = clean_amount(row[COL['mensualite']])
        transport_amt, has_transport = clean_transport(row[COL['transport']])
        insc_raw   = str(row[COL['inscription']] or '').strip()
        insc_amt   = clean_amount(row[COL['inscription']])
        note_raw   = row[COL['note']]
        reg_date   = clean_date(note_raw)

        # Validation
        if not last_name:
            stats['errors'] += 1
            continue

        # Unique code
        code = f"STU-{matricule.zfill(4)}"

        # ── Duplicate check ───────────────────────────────────────
        # Check by code first, then by mat+name combo
        existing = session.query(Student).filter_by(code=code).first()
        if not existing:
            # Also try by name+class (different matricule but same person)
            existing = session.query(Student).filter_by(
                last_name=last_name, class_name=classe, active=True
            ).filter(Student.first_name.ilike(f'%{first_name[:3]}%')).first()

        if existing:
            stats['skipped'] += 1
            student = existing
        else:
            # Create student
            student = Student(
                code=code,
                first_name=first_name,
                last_name=last_name,
                class_name=classe,
                monthly_fee=mensualite,
                has_transport=has_transport,
                transport_fee=transport_amt,
                insurance_amount=insc_amt,
                insurance_paid=False,
                reinscription_status='pending',
                registration_date=reg_date or date(2024, 9, 1),
                notes=f"MAT:{matricule}",  # store original matricule
                active=True,
            )
            # Special: GRATUIT or 0 inscription means free
            if insc_raw.upper() == 'GRATUIT':
                student.notes = (student.notes or '') + '|GRATUIT'

            session.add(student)
            try:
                session.flush()
                stats['inserted'] += 1
            except Exception as e:
                session.rollback()
                stats['errors'] += 1
                print(f"  ⚠ Error inserting {last_name} {first_name}: {e}")
                continue

        # ── Parse & insert month records ──────────────────────────
        month_cells = excel_row[9:19]   # 10 months Sep–Jun

        for m_idx, (month_name, cell) in enumerate(zip(SCHOOL_MONTHS_10, month_cells)):
            cell_val = cell.value
            # Detect yellow NAN
            try:
                rgb = cell.fill.fgColor.rgb
                is_yellow = (rgb == 'FFFFFF00')
            except Exception:
                is_yellow = False

            status = parse_month_status(cell_val, is_yellow)

            # Check for existing MonthRecord
            existing_mr = session.query(MonthRecord).filter_by(
                student_id=student.id,
                month_name=month_name,
                school_year=school_year
            ).first()

            if existing_mr:
                # Update if different
                if existing_mr.status != status:
                    existing_mr.status = status
                    if status != 'nan':
                        existing_mr.amount = mensualite
            else:
                mr = MonthRecord(
                    student_id=student.id,
                    month_name=month_name,
                    school_year=school_year,
                    status=status,
                    amount=mensualite if status != 'nan' else 0.0,
                )
                session.add(mr)

            # Update stats
            if status == 'paid':
                stats['paid_months'] += 1
            elif status == 'nan':
                stats['nan_months'] += 1
            else:
                stats['unpaid_months'] += 1

        # ── Create Payment records for paid months ────────────────
        if stats['inserted'] > 0 or True:  # Always process payments for accurate history
            year = 2025  # school year 2024-25
            for m_idx, (month_name, cell) in enumerate(zip(SCHOOL_MONTHS_10, month_cells)):
                cell_val = cell.value
                try:
                    is_yellow = (cell.fill.fgColor.rgb == 'FFFFFF00')
                except:
                    is_yellow = False
                status = parse_month_status(cell_val, is_yellow)

                if status == 'paid':
                    # Check if payment already exists
                    existing_pay = session.query(Payment).filter_by(
                        student_id=student.id,
                        payment_type='monthly',
                        month=month_name,
                        year=year
                    ).first()
                    if not existing_pay and mensualite > 0:
                        rec_count = session.query(Payment).count()
                        rec_num = f'REC-{year}-{rec_count+1:06d}'
                        pay = Payment(
                            student_id=student.id,
                            payment_type='monthly',
                            amount=mensualite,
                            month=month_name,
                            year=year,
                            payment_date=datetime(year, m_idx+9 if m_idx+9 <= 12 else m_idx-3, 1),
                            receipt_number=rec_num,
                        )
                        session.add(pay)
                        stats['payments_created'] += 1

        # Commit every 50 rows to avoid memory issues
        if (stats['inserted'] + stats['skipped']) % 50 == 0:
            session.commit()

    session.commit()
    print(f"  ✅ Students inserted: {stats['inserted']}")
    print(f"  ⏭  Students skipped (duplicates): {stats['skipped']}")
    print(f"  ❌ Errors: {stats['errors']}")
    print(f"  ✅ Month records — Paid: {stats['paid_months']}, NAN: {stats['nan_months']}, Unpaid: {stats['unpaid_months']}")
    print(f"  💳 Payment records created: {stats['payments_created']}")

    # ── Step 5: Final DB stats ────────────────────────────────────
    print("\n[5/6] Database summary...")
    n_students = session.query(Student).filter_by(active=True).count()
    n_months   = session.query(MonthRecord).count()
    n_paid_mr  = session.query(MonthRecord).filter_by(status='paid').count()
    n_nan_mr   = session.query(MonthRecord).filter_by(status='nan').count()
    n_unpaid_mr= session.query(MonthRecord).filter_by(status='unpaid').count()
    n_payments = session.query(Payment).count()

    print(f"  👥 Total students: {n_students}")
    print(f"  📅 Month records:  {n_months}")
    print(f"     ✅ Paid:    {n_paid_mr}")
    print(f"     ⊘  NAN:     {n_nan_mr}")
    print(f"     ⏳ Unpaid:  {n_unpaid_mr}")
    print(f"  💳 Payments:   {n_payments}")

    # Class breakdown
    print("\n  Classes:")
    from sqlalchemy import func
    class_counts = session.query(Student.class_name, func.count(Student.id)).filter_by(active=True).group_by(Student.class_name).all()
    for cls, cnt in sorted(class_counts, key=lambda x: (x[0] or '')):
        print(f"     {cls or 'Unknown':12s}: {cnt}")

    # ── Step 6: Verify data integrity ────────────────────────────
    print("\n[6/6] Data integrity checks...")
    
    # Check every student has 10 month records
    bad = 0
    for s in session.query(Student).filter_by(active=True).all():
        cnt = session.query(MonthRecord).filter_by(student_id=s.id, school_year=school_year).count()
        if cnt != 10:
            bad += 1
            print(f"  ⚠ {s.code} {s.last_name}: {cnt}/10 month records")
    if bad == 0:
        print(f"  ✅ All students have exactly 10 month records")

    # Check no payment duplicates
    from sqlalchemy import text
    dup_result = session.execute(text("""
        SELECT student_id, month, year, COUNT(*) as cnt
        FROM payments
        WHERE payment_type = 'monthly'
        GROUP BY student_id, month, year
        HAVING cnt > 1
    """)).fetchall()
    if dup_result:
        print(f"  ⚠ {len(dup_result)} duplicate payment records found!")
    else:
        print("  ✅ No duplicate payment records")

    session.close()
    print("\n" + "=" * 60)
    print("  ✅ IMPORT COMPLETE — Application ready to use!")
    print("=" * 60)
    print(f"\n  Run: python main.py")
    print(f"  Login: admin / admin123\n")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Import sch.xlsx into Le Schéma SGS')
    parser.add_argument('--xlsx', default='/mnt/user-data/uploads/sch.xlsx',
                        help='Path to Excel file')
    parser.add_argument('--clear', action='store_true',
                        help='Clear existing student data before import')
    parser.add_argument('--year', default='2024-25',
                        help='School year (default: 2024-25)')
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"Error: File not found: {args.xlsx}")
        sys.exit(1)

    run_import(args.xlsx, clear_existing=args.clear, school_year=args.year)
